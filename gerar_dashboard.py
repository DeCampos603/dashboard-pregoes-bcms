# -*- coding: utf-8 -*-
"""Dashboard "Capacidade de Empenho" — Atas de Registro de Preços (BCMS).

Lê o consolidado do Robô Extrator de Pregões (Consolidado_Pregoes.xlsx, nome
fixo no Google Drive → link estável), calcula a capacidade de empenho
(Qtd. Saldo × Valor Unitário das atas vigentes, nas linhas da UASG alvo) e gera
um site estático autocontido em site/index.html para o GitHub Pages.

Uso:
  py -3 gerar_dashboard.py                       # baixa do Drive (DRIVE_FILE_ID)
  SOURCE_XLSX=caminho.xlsx py -3 gerar_dashboard.py   # usa arquivo local (teste)

Ambiente:
  DRIVE_FILE_ID  ID do arquivo no Drive (ou usa o padrão embutido)
  SOURCE_XLSX    caminho local (tem precedência; para teste)
  UASG_ALVO      prefixo da UASG participante (padrão: 160329 = BCMS)
"""

from __future__ import annotations

import html
import io
import json
import os
import re
import unicodedata
import urllib.request
from datetime import datetime, timedelta

from openpyxl import load_workbook

# ---------------------------------------------------------------- configuração
BASE = os.path.dirname(os.path.abspath(__file__))
SITE = os.path.join(BASE, "site")
DATA = os.path.join(BASE, "data")

# ID do Consolidado_Pregoes.xlsx no Google Drive (Compartilhar → qualquer
# pessoa com o link). Pode ser sobrescrito pelo secret/variável DRIVE_FILE_ID.
DRIVE_FILE_ID_PADRAO = "1Jp2lSSR1m_FofvlwJbcOUT9wCXxG6785"

UASG_ALVO = os.environ.get("UASG_ALVO", "160329").strip()
NOME_UNIDADE = os.environ.get("NOME_UNIDADE", "Batalhão Central de Manutenção e Suprimento")
NOME_CURTO = os.environ.get("NOME_CURTO", "BCMS")

# Unidade "primária" — a que vira index.html no site multi-unidade. As demais
# viram <slug>.html. `main()` gera uma página por OMDS com dados (looping por
# UNIDADES e reatribuindo UASG_ALVO/NOME_* a cada volta).
UASG_PRIMARIO = UASG_ALVO

AUSENTE = "Informação ausente"

# ---------------------------------------------------------------- OMDS (unidades)
# Manifesto das OM Diretamente Subordinadas. O brasão (logo) é a identidade que
# troca ao clicar; o `accent` (cor derivada do logo, já validada p/ contraste)
# tinge apenas o "cromo" (barra de topo, chip ativo, moldura) — gráficos e
# status permanecem na paleta medida, para não colidir com o vermelho de
# "Vencida". Uma unidade só é considerada COLETADA se for a UASG_ALVO desta
# geração (é a única com dados reais); as demais mostram "Aguardando coleta".
#
# ⚠️ Confirme os nomes oficiais de BMSA e ECT.
UNIDADES = [
    {"sigla": "BCMS",    "nome": "Batalhão Central de Manutenção e Suprimento",
     "uasg": "160329", "logo": "BCMS.png",  "accent": "#DB2819"},
    {"sigla": "Ba Ap Log", "nome": "Base de Apoio Logístico do Exército",
     "uasg": "160238", "logo": "BaApLog.png", "accent": "#D83030"},
    {"sigla": "1º D Sup", "nome": "1º Depósito de Suprimento",
     "uasg": "160307", "logo": "1DSUP.png", "accent": "#DE2B30"},
    {"sigla": "BMSA",    "nome": "BMSA",
     "uasg": "160304", "logo": "BMSA.png",  "accent": "#DB2819"},
    {"sigla": "D C Mun", "nome": "Depósito Central de Munição",
     "uasg": "160246", "logo": "DCMUN.png", "accent": "#047CC0"},
    {"sigla": "ECT",     "nome": "ECT",
     "uasg": "160321", "logo": "Ect.png",   "accent": "#B33338"},
]

# UASG → sigla das OMDS que conhecemos (o consolidado só traz o número da UASG
# gerenciadora, sem nome). Só nomeamos as unidades do manifesto; as demais
# gerenciadoras (dezenas de UASGs externas) aparecem só com o número.
NOME_POR_UASG = {u["uasg"]: u["sigla"] for u in UNIDADES}
OM_POR_UASG = {u["uasg"]: u for u in UNIDADES}


def unidade_ativa() -> dict:
    """A unidade cujos dados esta geração contém (a UASG_ALVO)."""
    for u in UNIDADES:
        if u["uasg"] == UASG_ALVO:
            return u
    return {"sigla": NOME_CURTO, "nome": NOME_UNIDADE, "uasg": UASG_ALVO,
            "logo": "BCMS.png", "accent": "#DB2819"}


def unidades_json() -> str:
    """Manifesto para o JS, marcando qual está coletada (a UASG_ALVO)."""
    saida = [dict(u, coletado=(u["uasg"] == UASG_ALVO)) for u in UNIDADES]
    return json.dumps(saida, ensure_ascii=False)


def _slug(u: dict) -> str:
    """Nome de arquivo da página da unidade (sem acento/espaço)."""
    s = unicodedata.normalize("NFKD", u["sigla"]).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")
    return s or u["uasg"]


def pagina_da_unidade(u: dict) -> str:
    """Arquivo HTML da unidade: a primária é index.html; as demais, <slug>.html."""
    return "index.html" if u["uasg"] == UASG_PRIMARIO else f"{_slug(u)}.html"


def omds_nav_html(com_dados: set, atual_uasg: str) -> str:
    """Barra de OMDS: cada unidade COM dados vira um link para sua página; as sem
    dados ficam como chip apagado ('aguardando coleta'). Navegação por link —
    cada página já vem pronta (tema/brasão/dados no servidor), sem troca por JS."""
    chips = []
    for u in UNIDADES:
        tem = u["uasg"] in com_dados
        eh_atual = u["uasg"] == atual_uasg
        logo = (f'<img src="assets/logos/{esc(u["logo"])}" alt="" loading="lazy" '
                f'onerror="this.style.display=\'none\'"><span>{esc(u["sigla"])}</span>')
        if tem:
            chips.append(
                f'<a class="omds" href="{pagina_da_unidade(u)}" '
                f'aria-current="{"true" if eh_atual else "false"}" '
                f'title="{esc(u["nome"])}">{logo}</a>')
        else:
            chips.append(
                f'<span class="omds omds-off" aria-current="false" '
                f'title="{esc(u["nome"])} — aguardando coleta">{logo}'
                f'<span class="omds-selo" title="Aguardando coleta">•</span></span>')
    return "".join(chips)


# ---------------------------------------------------------------- utilidades
def num(v):
    """Converte célula em float (aceita '1.234,56'); None se não numérico."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s == AUSENTE:
        return None
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def data_br(v):
    """'dd/mm/aaaa' (ou datetime) → datetime; None se inválida."""
    if isinstance(v, datetime):
        return v
    try:
        return datetime.strptime(str(v).strip(), "%d/%m/%Y")
    except (ValueError, TypeError):
        return None


def fmt_brl(v) -> str:
    if v is None:
        return "—"
    s = f"{v:,.2f}".replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {s}"


def fmt_short(v) -> str:
    """R$ compacto para rótulos de barra (mi/mil)."""
    if v is None:
        return "—"
    if abs(v) >= 1_000_000:
        return f"R$ {v/1_000_000:.2f} mi".replace(".", ",")
    if abs(v) >= 1_000:
        return f"R$ {v/1_000:.0f} mil"
    return fmt_brl(v)


def fmt_int(v) -> str:
    return f"{int(v):,}".replace(",", ".")


def esc(s, limite=None) -> str:
    t = str(s or "").strip()
    if limite and len(t) > limite:
        t = t[: limite - 1] + "…"
    return html.escape(t, quote=True)


# ---------------------------------------------------------------- fonte de dados
def baixar_do_drive(file_id: str) -> io.BytesIO:
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    print(f"[INFO] Baixando consolidado do Drive ({file_id[:8]}…)")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    dados = urllib.request.urlopen(req, timeout=120).read()
    if not dados.startswith(b"PK"):
        # Arquivos grandes ganham página interstitial; tenta o confirm direto.
        req2 = urllib.request.Request(url + "&confirm=t", headers={"User-Agent": "Mozilla/5.0"})
        dados = urllib.request.urlopen(req2, timeout=120).read()
    if not dados.startswith(b"PK"):
        raise RuntimeError(
            "O download não devolveu um .xlsx — confira se o arquivo está "
            "compartilhado como 'Qualquer pessoa com o link' e se o ID está certo.")
    print(f"[INFO] {len(dados)/1024:.0f} KB baixados.")
    return io.BytesIO(dados)


def carregar_linhas(fobj) -> list[dict]:
    wb = load_workbook(fobj, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas, cab = [], None
    for row in ws.iter_rows(values_only=True):
        if cab is None:
            cab = [str(c or "").strip() for c in row]
            continue
        linhas.append(dict(zip(cab, row)))
    wb.close()
    print(f"[INFO] {len(linhas)} linhas lidas do consolidado.")
    return linhas


# ---------------------------------------------------------------- ETL
def classificar_status(r: dict, hoje: datetime) -> str:
    """Bucket do status: vig | v30 | venc | semata | semdados."""
    s = str(r.get("Status_Ata") or "").strip()
    if s.startswith("Sem ata"):
        return "semata"
    if s == "Sem dados":
        return "semdados"
    if s == "Vencida":
        return "venc"
    if s.startswith("Vence em"):
        return "v30"
    if s.startswith("Vigente"):
        return "vig"
    # Sem coluna Status_Ata (planilha antiga): deriva das datas.
    if str(r.get("Nr Ata") or "").strip() in ("", AUSENTE):
        forn = str(r.get("Fornecedor") or "").strip()
        return "semdados" if forn in ("", AUSENTE) else "semata"
    fim = data_br(r.get("Fim Vig Ata"))
    if fim is None:
        return "vig"
    if fim < hoje:
        return "venc"
    if fim < hoje + timedelta(days=30):
        return "v30"
    return "vig"


def etl(linhas: list[dict]) -> dict:
    hoje = datetime.now()
    itens = []          # linhas da UASG alvo com saldo calculável
    datas_coleta = []
    for r in linhas:
        if not str(r.get("UASG") or "").strip().startswith(UASG_ALVO):
            continue
        d = data_br(r.get("Data_Coleta"))
        if d:
            datas_coleta.append(d)
        saldo, vu = num(r.get("Qtd. Saldo")), num(r.get("Val. Unitário"))
        cap = round(saldo * vu, 2) if (saldo is not None and vu is not None) else None
        pregao = str(r.get("Pregão") or "").replace("_", "/")
        ger = str(r.get("UASG_Gerenciadora") or "")
        itens.append({
            "pregao": pregao,
            "ger": ger,
            "key": f"{ger} · {pregao}",   # identifica o pregão de forma única (ger+nº)
            "nr": r.get("Nr Item"),
            "desc": str(r.get("Descrição detalhada") or ""),
            "forn": str(r.get("Fornecedor") or ""),
            "cat": str(r.get("Categoria_Geral") or "").strip() or "Sem categoria",
            "nd": str(r.get("ND_Sugerida") or "").strip() or "—",
            "sub": str(r.get("Subitem_Sugerido") or "").strip(),
            "tipo": str(r.get("Catalogo_Tipo") or "").strip() or "—",
            "fim": data_br(r.get("Fim Vig Ata")),
            "st": classificar_status(r, hoje),
            "saldo": saldo, "vu": vu, "cap": cap,
            "link": str(r.get("Link_Item") or ""),
            # campos extras usados só no detalhamento (modal ao clicar)
            "nrata": str(r.get("Nr Ata") or "").strip(),
            "catmat": str(r.get("CATMAT_CATSER_Nome") or "").strip(),
            "cod": str(r.get("CATMAT_CATSER_Codigo") or "").strip(),
            "conf": str(r.get("Confianca_ND") or "").strip(),
            "ini": data_br(r.get("Início Vig Ata")),
            "vtot": num(r.get("Valor_Total_Homologado")),
            "qhom": num(r.get("Qtd Homologada")),
            "qaut": num(r.get("Qtd. Autorizada")),
            "stlabel": str(r.get("Status_Ata") or "").strip(),
        })

    ativos = [i for i in itens if i["st"] in ("vig", "v30") and (i["cap"] or 0) > 0]
    cap_vig = sum(i["cap"] for i in ativos if i["st"] == "vig")
    cap_v30 = sum(i["cap"] for i in ativos if i["st"] == "v30")
    cap_venc = sum(i["cap"] or 0 for i in itens if i["st"] == "venc" and (i["cap"] or 0) > 0)

    por_cat, por_pregao = {}, {}
    for i in ativos:
        por_cat[i["cat"]] = por_cat.get(i["cat"], 0) + i["cap"]
        chave = f"{i['ger']} · {i['pregao']}"
        por_pregao[chave] = por_pregao.get(chave, 0) + i["cap"]

    # Opções dos filtros, ordenadas por capacidade (as maiores primeiro), com
    # o valor e a contagem em cada opção — o usuário escolhe já vendo o peso.
    def opcoes(chave, rotulo=None):
        agg = {}
        for i in ativos:
            a = agg.setdefault(i[chave], {"cap": 0, "n": 0,
                                          "lab": rotulo(i) if rotulo else i[chave]})
            a["cap"] += i["cap"]
            a["n"] += 1
        return sorted(agg.items(), key=lambda x: -x[1]["cap"])

    # Vencimentos ≤60 dias, agrupados por pregão+data.
    venc60 = {}
    for i in ativos:
        if i["fim"] and i["fim"] < hoje + timedelta(days=60):
            k = (i["fim"], i["pregao"], i["ger"])
            v = venc60.setdefault(k, {"cap": 0, "n": 0})
            v["cap"] += i["cap"]
            v["n"] += 1

    # PLANEJAMENTO: pregões cujas atas ainda VALEM (vigentes ou vencendo ≤30d),
    # para o setor programar novas licitações antes de expirarem. Agrupa por
    # pregão (ger · nº); o "objeto" é derivado das categorias dos itens (a coluna
    # Objeto do consolidado é só "x"). Atas JÁ VENCIDAS ficam de fora — não há o
    # que planejar nelas, já expiraram.
    planej = {}
    for i in itens:
        # só atas com data de fim e que AINDA não venceram (pela data real, não
        # pelo Status_Ata que o robô congelou na coleta — entre a coleta e a
        # geração alguma pode ter expirado). Vencidas ficam totalmente de fora.
        if i["st"] not in ("vig", "v30", "venc") or not i["fim"]:
            continue
        if i["fim"].date() < hoje.date():
            continue
        p = planej.get(i["key"])
        if p is None:
            p = planej[i["key"]] = {"key": i["key"], "ger": i["ger"], "pregao": i["pregao"],
                                    "fim": i["fim"], "cap": 0.0, "n": 0, "cats": {}, "forns": set()}
        if i["fim"] < p["fim"]:
            p["fim"] = i["fim"]          # o vencimento mais próximo do pregão
        c = i["cap"] if isinstance(i["cap"], (int, float)) else 0
        p["cap"] += c
        p["n"] += 1
        # objeto = categorias por CONTAGEM de itens (sempre disponível, mesmo
        # quando a ata já foi 100% consumida — que é justo a que mais urge renovar)
        p["cats"][i["cat"]] = p["cats"].get(i["cat"], 0) + 1
        if str(i["forn"]).strip() not in ("", AUSENTE):
            p["forns"].add(i["forn"])
    planejamento = sorted(planej.values(), key=lambda p: p["fim"])

    return {
        "hoje": hoje,
        "posicao": max(datas_coleta).strftime("%d/%m/%Y") if datas_coleta else hoje.strftime("%d/%m/%Y"),
        "itens": itens, "ativos": ativos,
        "cap_total": cap_vig + cap_v30, "cap_vig": cap_vig, "cap_v30": cap_v30,
        "cap_venc": cap_venc,
        "n_itens": len(ativos),
        "n_pregoes": len({(i["ger"], i["pregao"]) for i in ativos}),
        "por_cat": sorted(por_cat.items(), key=lambda x: -x[1])[:12],
        "por_pregao": sorted(por_pregao.items(), key=lambda x: -x[1])[:12],
        "venc60": sorted(venc60.items())[:12],
        "op_cat": opcoes("cat"),
        "op_nd": opcoes("nd", lambda i: (f"{i['nd']} · {i['sub'].split(' - ', 1)[-1].title()}"
                                         if i["sub"] else i["nd"])),
        "op_pregao": opcoes("key"),   # valor = "ger · nº" (pregão preciso, com a UASG)
        "op_tipo": opcoes("tipo"),
        # Fornecedor (empresa): valor = string completa "CNPJ - NOME" (dá pra
        # buscar por CNPJ ou nome). Descarta ausentes.
        "op_forn": [o for o in opcoes("forn")
                    if str(o[0]).strip() not in ("", AUSENTE)],
        "planejamento": planejamento,
    }


# ---------------------------------------------------------------- histórico
def atualizar_historico(m: dict) -> list[dict]:
    os.makedirs(DATA, exist_ok=True)
    caminho = os.path.join(DATA, "history.json")
    hist = []
    if os.path.exists(caminho):
        try:
            with open(caminho, "r", encoding="utf-8") as f:
                hist = json.load(f)
        except (json.JSONDecodeError, OSError):
            hist = []
    chave = m["posicao"]  # 1 snapshot por posição de coleta
    hist = [h for h in hist if h.get("posicao") != chave]
    hist.append({"posicao": chave, "cap_total": round(m["cap_total"], 2),
                 "cap_v30": round(m["cap_v30"], 2), "itens": m["n_itens"]})
    hist = hist[-120:]
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=1)
    return hist


def sparkline(hist: list[dict]) -> str:
    if len(hist) < 2:
        return '<p class="muted">O histórico acumula a partir da 2ª atualização.</p>'
    vals = [h["cap_total"] for h in hist]
    vmin, vmax = min(vals), max(vals)
    faixa = (vmax - vmin) or 1
    W, H = 560, 90
    pts = []
    for i, v in enumerate(vals):
        x = 8 + i * (W - 16) / (len(vals) - 1)
        y = H - 12 - (v - vmin) / faixa * (H - 28)
        pts.append(f"{x:.1f},{y:.1f}")
    return (f'<svg viewBox="0 0 {W} {H}" role="img" aria-label="Evolução da capacidade" '
            f'preserveAspectRatio="none" style="width:100%;height:{H}px">'
            f'<polyline points="{" ".join(pts)}" fill="none" stroke="var(--primary)" '
            f'stroke-width="2.5" stroke-linejoin="round" stroke-linecap="round"/></svg>'
            f'<div class="spark-leg"><span>{hist[0]["posicao"]} · {fmt_short(vals[0])}</span>'
            f'<span>{hist[-1]["posicao"]} · {fmt_short(vals[-1])}</span></div>')


# ---------------------------------------------------------------- render
_ST_LABEL = {"vig": "Vigente", "v30": "≤30 dias", "venc": "Vencida",
             "semata": "Sem ata", "semdados": "Sem dados"}
# Ícone = 2º canal além da cor (exigido para daltonismo / impressão P&B).
_ST_ICONE = {"vig": "●", "v30": "▲", "venc": "✕", "semata": "○", "semdados": "–"}


def _barras(pares: list) -> str:
    """Barras horizontais: série única (uma cor), ponta arredondada 4px,
    base reta, trilho = passo claro do mesmo matiz."""
    if not pares:
        return '<p class="vazio">Sem dados para este filtro.</p>'
    topo = pares[0][1] or 1
    out = []
    for rotulo, valor in pares:
        pct = max(valor / topo * 100, 1.5)
        out.append(
            f'<div class="brow" tabindex="0" data-lab="{esc(rotulo)}" data-val="{fmt_brl(valor)}">'
            f'<div class="blabel">{esc(rotulo, 30)}</div>'
            f'<div class="btrack"><div class="bfill" style="width:{pct:.1f}%"></div></div>'
            f'<div class="bval">{fmt_short(valor)}</div></div>')
    return "".join(out)


def _select(id_, rotulo, opcoes, todos_label) -> str:
    """<select> de filtro, com capacidade e contagem visíveis em cada opção."""
    ops = [f'<option value="">{esc(todos_label)}</option>']
    for valor, a in opcoes:
        ops.append(f'<option value="{esc(valor)}">{esc(a["lab"], 46)} — '
                   f'{fmt_short(a["cap"])} ({a["n"]})</option>')
    return (f'<label class="fl"><span>{esc(rotulo)}</span>'
            f'<select id="{id_}">{"".join(ops)}</select></label>')


def _tabela(itens: list[dict]) -> str:
    linhas = []
    # Renderiza TODOS os itens (inclui saldo 0 / sem dados) para que o botão
    # "Mostrar todos os itens do pregão" possa exibir o pregão completo. O JS
    # esconde os itens sem saldo na visão de capacidade (padrão).
    visiveis = sorted(itens, key=lambda i: -(i["cap"] or 0))
    for i in visiveis:
        fim = i["fim"].strftime("%d/%m/%Y") if i["fim"] else "—"
        ts = f'{i["fim"].timestamp():.0f}' if i["fim"] else "0"
        ini = i["ini"].strftime("%d/%m/%Y") if i.get("ini") else ""
        link = (f'<a href="{esc(i["link"])}" target="_blank" rel="noopener" '
                f'aria-label="Abrir item no Compras.gov">↗</a>') if i["link"].startswith("http") else ""
        linhas.append(
            f'<tr data-st="{i["st"]}" data-cat="{esc(i["cat"])}" data-nd="{esc(i["nd"])}"'
            f' data-pg="{esc(i["pregao"])}" data-tipo="{esc(i["tipo"])}"'
            f' data-forn="{esc(i["forn"])}"'
            f' data-cap="{i["cap"] or 0}" data-saldo="{i["saldo"] or 0}" data-fimts="{ts}"'
            f' data-key="{esc(i["key"])}"'
            f' data-desc="{esc(i["desc"], 400)}" data-catmat="{esc(i["catmat"], 200)}"'
            f' data-cod="{esc(i["cod"])}" data-sub="{esc(i["sub"], 120)}"'
            f' data-conf="{esc(i["conf"])}" data-ini="{ini}" data-fim="{fim}"'
            f' data-vu="{i["vu"] or 0}" data-vtot="{i["vtot"] or 0}"'
            f' data-qhom="{i["qhom"] or 0}" data-qaut="{i["qaut"] or 0}"'
            f' data-nrata="{esc(i["nrata"])}" data-stlabel="{esc(i["stlabel"])}"'
            f' data-link="{esc(i["link"])}" tabindex="0"'
            f' aria-label="Ver detalhes do item {esc(str(i["nr"]))}">'
            f'<td>{esc(i["pregao"])}</td><td>{esc(i["ger"])}</td>'
            f'<td class="num">{esc(i["nr"])}</td>'
            f'<td class="desc" title="{esc(i["desc"], 300)}">{esc(i["desc"], 90)}</td>'
            f'<td class="cat">{esc(i["cat"], 26)}</td>'
            f'<td class="nd" title="{esc(i["sub"])}">{esc(i["nd"])}</td>'
            f'<td class="forn" title="{esc(i["forn"], 160)}">{esc(i["forn"], 34)}</td>'
            f'<td data-v="{ts}">{fim}</td>'
            f'<td><span class="tag tag-{i["st"]}">'
            f'<span aria-hidden="true">{_ST_ICONE[i["st"]]}</span>{_ST_LABEL[i["st"]]}</span></td>'
            f'<td class="num" data-v="{i["saldo"] or 0}">{fmt_int(i["saldo"] or 0)}</td>'
            f'<td class="num" data-v="{i["vu"] or 0}">{fmt_brl(i["vu"])}</td>'
            f'<td class="num" data-v="{i["cap"] or 0}"><strong>{fmt_brl(i["cap"])}</strong></td>'
            f'<td>{link}</td></tr>')
    return "\n".join(linhas)


def _objeto_do_pregao(p: dict) -> str:
    """'Objeto' derivado das categorias dos itens (a coluna Objeto é só 'x').
    Categoria ausente vai para o fim, para nunca roubar a vez de uma real."""
    cats = sorted(p["cats"].items(), key=lambda x: (x[0] == AUSENTE, -x[1]))
    if not cats:
        return "—"
    nomes = [c[0] for c in cats[:2]]
    if len(cats) > 2:
        nomes.append(f"+{len(cats) - 2}")
    return " · ".join(nomes)


def _tabela_planejamento(planejamento: list) -> str:
    """Linhas da aba de planejamento: um pregão por linha, ordenado por
    vencimento. data-fimts alimenta o cálculo de dias no cliente."""
    linhas = []
    for p in planejamento:
        ts = f'{p["fim"].timestamp():.0f}'
        fim = p["fim"].strftime("%d/%m/%Y")
        objeto = _objeto_do_pregao(p)
        forns = sorted(p["forns"])
        forn_lbl = ("—" if not forns else
                    forns[0].split(" - ", 1)[-1] if len(forns) == 1 else
                    f"{len(forns)} fornecedores")
        forn_tt = "; ".join(f.split(" - ", 1)[-1] for f in forns[:8])
        # papel do BCMS: gerenciador quando a UASG gerenciadora é a própria
        # UASG alvo; senão é participante (carona) de pregão de outra OM.
        gerenc = str(p["ger"]).startswith(UASG_ALVO)
        papel = "ger" if gerenc else "part"
        papel_lbl = "gerencia" if gerenc else "participa"
        # nome da unidade gerenciadora (só as OMDS conhecidas; demais só o nº)
        omd = OM_POR_UASG.get(str(p["ger"]).strip())
        sigla = omd["sigla"] if omd else ""
        nome_om = omd["nome"] if omd else ""
        om = f'<span class="pl-om" title="{esc(nome_om)}">{esc(sigla)}</span>' if sigla else ""
        ger_full = f'{p["ger"]} · {sigla}' if sigla else str(p["ger"])
        forns_all = "; ".join(f.split(" - ", 1)[-1] for f in forns)
        linhas.append(
            f'<tr data-fimts="{ts}" data-key="{esc(p["key"])}" data-papel="{papel}"'
            f' data-ger="{esc(str(p["ger"]))}" data-om="{esc(sigla)}"'
            f' data-omnome="{esc(nome_om)}" data-obj="{esc(objeto, 120)}"'
            f' data-forns="{esc(forns_all, 600)}" tabindex="0"'
            f' aria-label="Ver detalhes do pregão {esc(p["pregao"])}">'
            f'<td data-v="{ts}" class="pl-fim"><b>{fim}</b><span class="pl-dias" '
            f'data-fimts="{ts}"></span></td>'
            f'<td>{esc(p["pregao"])}</td>'
            f'<td class="pl-ger" data-ger="{esc(ger_full)}" title="{esc(nome_om)}">'
            f'<b>{esc(p["ger"])}</b>{om}'
            f'<span class="pl-papel p-{papel}">{papel_lbl}</span></td>'
            f'<td class="pl-obj" title="{esc(objeto, 120)}">{esc(objeto, 46)}</td>'
            f'<td class="num" data-v="{p["n"]}">{fmt_int(p["n"])}</td>'
            f'<td class="num" data-v="{p["cap"]:.0f}"><strong>{fmt_brl(p["cap"])}</strong></td>'
            f'<td class="forn" title="{esc(forn_tt, 240)}">{esc(forn_lbl, 34)}</td>'
            f'</tr>')
    return "\n".join(linhas)


def render(m: dict, hist: list[dict], com_dados: set) -> str:
    tpl = _TEMPLATE
    ativa = unidade_ativa()
    subs = {
        "%%POSICAO%%": m["posicao"],
        "%%GERADO%%": m["hoje"].strftime("%d/%m/%Y %H:%M"),
        "%%UASG%%": ativa["uasg"],
        "%%UNIDADE%%": ativa["nome"],
        "%%UNIDADE_CURTA%%": ativa["sigla"],
        "%%EMBLEMA%%": f'assets/logos/{ativa["logo"]}',
        "%%ACCENT%%": ativa["accent"],
        "%%OMDS_NAV%%": omds_nav_html(com_dados, ativa["uasg"]),
        "%%TABELA_PLANEJ%%": _tabela_planejamento(m["planejamento"]),
        "%%N_PLANEJ%%": fmt_int(len(m["planejamento"])),
        "%%CAP_TOTAL%%": fmt_brl(m["cap_total"]),
        "%%CAP_VIG%%": fmt_brl(m["cap_vig"]),
        "%%CAP_V30%%": fmt_brl(m["cap_v30"]),
        "%%CAP_VENC%%": fmt_brl(m["cap_venc"]),
        "%%N_ITENS%%": fmt_int(m["n_itens"]),
        "%%N_PREGOES%%": fmt_int(m["n_pregoes"]),
        "%%PCT_V30%%": f"{(m['cap_v30'] / m['cap_total'] * 100 if m['cap_total'] else 0):.1f}%".replace(".", ","),
        "%%BARRAS_CAT%%": _barras(m["por_cat"]),
        "%%BARRAS_PREGAO%%": _barras(m["por_pregao"]),
        "%%PCT_VIG_M%%": f"{(m['cap_vig'] / m['cap_total'] * 100 if m['cap_total'] else 0):.2f}",
        "%%PCT_V30_M%%": f"{(m['cap_v30'] / m['cap_total'] * 100 if m['cap_total'] else 0):.2f}",
        "%%SPARK%%": sparkline(hist),
        "%%TABELA%%": _tabela(m["itens"]),
        "%%OMNOMES%%": json.dumps(
            {u["uasg"]: {"s": u["sigla"], "n": u["nome"]} for u in UNIDADES},
            ensure_ascii=False),
        "%%SEL_CAT%%": _select("fCat", "Tipo de material / serviço", m["op_cat"],
                               "Todas as categorias"),
        "%%SEL_ND%%": _select("fNd", "Natureza de despesa (sugerida)", m["op_nd"],
                              "Todas as ND"),
        "%%SEL_PG%%": _select("fPg", "Pregão", m["op_pregao"], "Todos os pregões"),
        "%%SEL_FORN%%": _select("fForn", "Empresa / fornecedor", m["op_forn"],
                                "Todas as empresas"),
        "%%SEL_TP%%": _select("fTp", "Material ou serviço", m["op_tipo"], "Ambos"),
        "%%VENC60%%": "".join(
            f'<tr><td>{fim.strftime("%d/%m/%Y")}</td><td>{esc(pregao)}</td>'
            f'<td>{esc(ger)}</td><td class="num">{v["n"]}</td>'
            f'<td class="num"><strong>{fmt_brl(v["cap"])}</strong></td></tr>'
            for (fim, pregao, ger), v in m["venc60"]) or
            '<tr><td colspan="5" class="muted">Nenhuma ata vencendo nos próximos 60 dias.</td></tr>',
    }
    for k, v in subs.items():
        tpl = tpl.replace(k, v)
    return tpl


_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR" data-theme="light" style="--accent:%%ACCENT%%">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="color-scheme" content="light dark">
<title>Capacidade de Empenho — Atas %%UNIDADE_CURTA%%</title>
<style>
/* ---------------------------------------------------------------- tokens
   Cores MEDIDAS (não estimadas): marcas >= 3:1 e textos >= 4.5:1 contra a
   superfície; trio de status com separação para daltonismo (OKLab dE >= 8). */
:root{
  --bg:#F1F4F8; --surface:#FFFFFF; --surface-2:#F7F9FC; --surface-3:#EDF1F6;
  --ink:#0F1B2A; --ink-2:#42505F; --ink-muted:#566374;
  --border:#DCE3EC; --border-2:#C6D0DC;
  --primary:#1C4A73;            /* marca (9,22:1) */
  --primary-ink:#164062;
  --accent:#DB2819;             /* cor da OMDS ativa (só no cromo); sobrescrita inline */
  --track:#D6E2EE;              /* trilho = passo claro do mesmo matiz */
  --success:#0F7A5A; --success-fill:#0F7A5A;   /* texto 5,31:1 */
  --warning:#8A631C; --warning-fill:#B5822B;   /* texto 5,41:1 */
  --danger:#B23A2E;                            /* texto 5,94:1 */
  --gold:#C8901E;               /* decorativo — NUNCA texto ou dado */
  --focus:#1C4A73;
  --shadow:0 1px 2px rgba(15,27,42,.05), 0 1px 3px rgba(15,27,42,.04);
  --shadow-2:0 4px 14px rgba(15,27,42,.08);
  --serif:Georgia,"Times New Roman",serif;
  --sans:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Helvetica Neue",Arial,sans-serif;
  --r:12px; --r-sm:8px;
}
[data-theme="dark"]{
  --bg:#0B1219; --surface:#131E2E; --surface-2:#182636; --surface-3:#1C2B3D;
  --ink:#E7EDF4; --ink-2:#C3CEDB; --ink-muted:#97A6B8;
  --border:#26374C; --border-2:#35495F;
  --primary:#4E86BD; --primary-ink:#8FB8DC; --track:#22405F;
  --success:#35C08F; --success-fill:#1F8959;
  --warning:#D9A94A; --warning-fill:#B98918;
  --danger:#E06A5E; --gold:#D9A94A; --focus:#8FB8DC;
  --shadow:0 1px 2px rgba(0,0,0,.35); --shadow-2:0 6px 18px rgba(0,0,0,.45);
}
*{box-sizing:border-box;margin:0}
html{scroll-behavior:smooth}
@media (prefers-reduced-motion:reduce){
  html{scroll-behavior:auto} *{transition:none!important;animation:none!important}
}
body{background:var(--bg);color:var(--ink);font:15px/1.55 var(--sans);
  -webkit-font-smoothing:antialiased;overflow-x:hidden}
:focus-visible{outline:2px solid var(--focus);outline-offset:2px;border-radius:4px}

/* ---------------------------------------------------------------- topo */
.bcms-bar{height:4px;background:var(--accent);transition:background .2s}
.hdr{position:sticky;top:0;z-index:40;background:color-mix(in srgb,var(--surface) 88%,transparent);
  backdrop-filter:saturate(1.6) blur(10px);border-bottom:1px solid var(--border)}
.hdr-in{max-width:1200px;margin:0 auto;padding:10px 20px;display:flex;
  justify-content:space-between;align-items:center;gap:14px;flex-wrap:wrap}
.hdr-l{display:flex;align-items:center;gap:12px;min-width:0}
.emblema{height:46px;width:46px;object-fit:contain;flex:none;
  filter:drop-shadow(0 1px 2px rgba(0,0,0,.18))}
h1{font-family:var(--serif);font-size:19px;font-weight:600;letter-spacing:-.01em;line-height:1.15}
.sub{font-size:12.5px;color:var(--ink-muted);margin-top:1px}
.hdr-r{display:flex;align-items:center;gap:10px}

/* barra de troca de OMDS */
.omds-nav{position:sticky;top:66px;z-index:35;background:var(--surface);
  border-bottom:1px solid var(--border);overflow-x:auto;scrollbar-width:thin}
@media(max-width:560px){.omds-nav{top:60px}}
.omds-nav-in{max-width:1200px;margin:0 auto;padding:8px 20px;display:flex;gap:8px}
.omds{display:inline-flex;align-items:center;gap:7px;flex:none;cursor:pointer;
  background:var(--surface-2);border:1px solid var(--border);border-radius:999px;
  padding:5px 13px 5px 6px;font:12.5px var(--sans);color:var(--ink-2);
  text-decoration:none;
  transition:background .15s,border-color .15s,color .15s}
a.omds:hover{border-color:var(--ink-muted);color:var(--ink)}
.omds-off{opacity:.5;cursor:default}
.omds-off:hover{border-color:var(--border);color:var(--ink-2)}
.omds img{height:24px;width:24px;object-fit:contain;border-radius:50%;
  background:#fff;padding:1px}
.omds[aria-current="true"]{background:var(--accent);border-color:var(--accent);
  color:#fff;font-weight:600}
.omds-selo{color:var(--ink-muted);font-size:16px;line-height:1;margin-left:-2px}
.omds[aria-current="true"] .omds-selo{color:rgba(255,255,255,.85)}

/* painel "aguardando coleta" */
.aguardando{display:none;text-align:center;padding:56px 24px}
.aguardando img{height:130px;width:130px;object-fit:contain;
  filter:drop-shadow(0 3px 8px rgba(0,0,0,.16));margin-bottom:18px}
.aguardando h2{font-family:var(--serif);font-size:22px;font-weight:600;margin-bottom:4px}
.aguardando .u-uasg{font-size:13px;color:var(--ink-muted);margin-bottom:18px}
.aguardando .u-badge{display:inline-block;background:var(--surface-3);
  border:1px solid var(--border);border-radius:999px;padding:8px 18px;
  font-size:13.5px;color:var(--ink-2)}
.aguardando .u-hint{margin-top:14px;font-size:12.5px;color:var(--ink-muted);max-width:440px;
  margin-left:auto;margin-right:auto}
.chip-pos{background:var(--surface-3);border:1px solid var(--border);border-radius:999px;
  padding:5px 12px;font-size:12px;color:var(--ink-2);white-space:nowrap}
.chip-pos b{color:var(--ink);font-variant-numeric:tabular-nums}
.btn{background:var(--surface);border:1px solid var(--border-2);border-radius:var(--r-sm);
  padding:7px 13px;cursor:pointer;color:var(--ink);font:13px var(--sans);
  transition:background .15s,border-color .15s}
.btn:hover{background:var(--surface-3);border-color:var(--ink-muted)}
.btn-pregao{border-color:var(--accent);color:var(--accent);font-weight:600}
.btn-pregao[aria-pressed="true"]{background:var(--accent);border-color:var(--accent);color:#fff}
[data-theme="dark"] .btn-pregao[aria-pressed="true"]{color:#0B1219}

.wrap{max-width:1200px;margin:0 auto;padding:22px 20px 72px}

/* abas (Capacidade / Planejamento) */
.tabs{display:flex;gap:4px;margin-bottom:18px;border-bottom:1px solid var(--border);
  overflow-x:auto;scrollbar-width:none}
.tab{background:none;border:0;border-bottom:2px solid transparent;padding:10px 16px 12px;
  cursor:pointer;font:14px var(--sans);color:var(--ink-muted);font-weight:600;
  white-space:nowrap;border-radius:8px 8px 0 0;transition:color .15s,border-color .15s,background .15s}
.tab:hover{color:var(--ink);background:var(--surface-2)}
.tab[aria-selected="true"]{color:var(--accent);border-bottom-color:var(--accent)}
/* aba de planejamento */
.pl-fim{white-space:nowrap}
.pl-dias{display:block;font-size:11px;font-weight:600;margin-top:1px}
.pl-dias.d-venc{color:var(--danger)}
.pl-dias.d-30{color:var(--warning)}
.pl-dias.d-ok{color:var(--ink-muted)}
.pl-obj{max-width:300px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:var(--ink-2)}
.pl-bandas{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:0 0 10px}
@media(max-width:640px){.pl-bandas{grid-template-columns:repeat(2,1fr)}}
.pl-card{display:flex;flex-direction:column;gap:3px;align-items:flex-start;text-align:left;
  background:var(--surface);border:1px solid var(--border-2);
  border-left:4px solid var(--bc,var(--ink-muted));border-radius:var(--r-sm);
  padding:11px 13px;cursor:pointer;transition:background .12s,box-shadow .12s;min-width:0;width:100%}
.pl-card:hover{background:var(--surface-3)}
.pl-card[aria-pressed="true"]{box-shadow:inset 0 0 0 2px var(--bc,var(--accent))}
.pl-card.b-30{--bc:var(--danger)}
.pl-card.b-90{--bc:var(--warning)}
.pl-card.b-180{--bc:color-mix(in srgb,var(--warning) 45%,var(--success))}
.pl-card.b-max{--bc:var(--success)}
.pl-card-n{font-size:26px;font-weight:800;line-height:1;font-variant-numeric:tabular-nums;color:var(--bc,var(--ink))}
.pl-card-lb{font-size:12px;color:var(--ink-muted);line-height:1.2}
.pl-card-cap{font-size:12.5px;font-weight:700;color:var(--ink-2);font-variant-numeric:tabular-nums}
.pl-bandas.tem-sel .pl-card:not([aria-pressed="true"]){opacity:.55}
.pl-ger{white-space:nowrap}
.pl-om{display:block;font-size:12px;font-weight:600;color:var(--ink-2);margin-top:1px}
.pl-papel{display:block;font-size:11px;font-weight:600;margin-top:1px}
.pl-papel.p-ger{color:var(--accent)}
.pl-papel.p-part{color:var(--ink-muted);font-weight:500}

/* ---------------------------------------------------------------- hero */
.hero{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);
  padding:26px 28px;box-shadow:var(--shadow);position:relative;overflow:hidden}
.hero::before{content:"";position:absolute;inset:0 auto 0 0;width:4px;background:var(--accent);transition:background .2s}
.hero-label{font-size:12.5px;text-transform:uppercase;letter-spacing:.08em;
  color:var(--ink-muted);font-weight:600}
/* Figura-herói: MESMA sans do resto (serifada seria decoração fora de marca)
   e figuras proporcionais (tabular deixa números grandes "soltos"). */
.hero-big{font-size:clamp(38px,6vw,54px);font-weight:700;letter-spacing:-.025em;
  line-height:1.05;margin:6px 0 4px;color:var(--ink);font-variant-numeric:proportional-nums}
.hero-cap{font-size:13px;color:var(--ink-muted);margin-bottom:16px}
/* Medidor de composição: 2px de superfície separando os segmentos (o branco
   separa, não uma borda desenhada). */
.meter{display:flex;gap:2px;height:12px;background:var(--track);
  border-radius:6px;overflow:hidden;max-width:640px}
.mseg{height:100%;transition:width .25s ease}
.mseg-vig{background:var(--success-fill)}
.mseg-v30{background:var(--warning-fill)}
.mleg{list-style:none;display:flex;gap:20px;flex-wrap:wrap;margin-top:10px;font-size:13px}
.mleg li{display:flex;align-items:center;gap:7px;color:var(--ink-2)}
.mleg b{color:var(--ink);font-variant-numeric:tabular-nums}
.key{width:10px;height:10px;border-radius:3px;flex:none}
.key-vig{background:var(--success-fill)} .key-v30{background:var(--warning-fill)}
.pills{margin-top:14px;display:none;align-items:center;gap:6px;flex-wrap:wrap}
.pills.on{display:flex}
.pills .lbl{font-size:12px;color:var(--ink-muted)}
.pill{background:var(--primary);color:#fff;border-radius:999px;padding:3px 11px;font-size:12px}
[data-theme="dark"] .pill{color:#0B1219}

/* ---------------------------------------------------------------- filtros */
.filtros{background:var(--surface);
  border:1px solid var(--border);border-radius:var(--r);padding:16px 18px;
  margin:16px 0;box-shadow:var(--shadow)}
.filtros-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:12px}
.fl{display:flex;flex-direction:column;gap:5px;min-width:0}
.fl>span{font-size:11.5px;font-weight:600;color:var(--ink-muted);
  text-transform:uppercase;letter-spacing:.04em}
.fl select,#busca,.cb-in{width:100%;background:var(--surface-2);color:var(--ink);
  border:1px solid var(--border-2);border-radius:var(--r-sm);padding:9px 11px;
  font:13.5px var(--sans);transition:border-color .15s}
.fl select:hover,#busca:hover,.cb-in:hover{border-color:var(--ink-muted)}
/* combobox pesquisável (input + lista filtrada) */
.cb{position:relative}
.cb-in{padding-right:30px;cursor:text;text-overflow:ellipsis}
.cb-in::placeholder{color:var(--ink-muted)}
.cb-clear{position:absolute;right:6px;top:50%;transform:translateY(-50%);
  width:22px;height:22px;border:0;background:none;color:var(--ink-muted);
  font-size:18px;line-height:1;cursor:pointer;border-radius:50%}
.cb-clear:hover{background:var(--surface-3);color:var(--ink)}
.cb-list{position:absolute;z-index:50;top:calc(100% + 4px);left:0;right:0;margin:0;
  padding:4px;list-style:none;max-height:280px;overflow-y:auto;
  background:var(--surface);border:1px solid var(--border-2);border-radius:var(--r-sm);
  box-shadow:var(--shadow-2)}
.cb-opt{padding:7px 9px;border-radius:6px;font-size:13px;cursor:pointer;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:var(--ink)}
.cb-opt:hover{background:var(--surface-3)}
.cb-opt[aria-selected="true"]{background:color-mix(in srgb,var(--accent) 16%,transparent);
  font-weight:600}
.cb-vazio{padding:8px 9px;font-size:12.5px;color:var(--ink-muted)}
.frow2{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-top:13px}
#busca{flex:1;min-width:200px}
/* Controle segmentado (status) — visualmente distinto dos dropdowns */
.seg{display:inline-flex;background:var(--surface-3);border:1px solid var(--border);
  border-radius:999px;padding:3px;gap:2px}
.seg button{background:none;border:0;border-radius:999px;padding:6px 14px;cursor:pointer;
  font:12.5px var(--sans);color:var(--ink-2);transition:background .15s,color .15s}
.seg button:hover{color:var(--ink)}
.seg button[aria-pressed="true"]{background:var(--primary);color:#fff;font-weight:600}
[data-theme="dark"] .seg button[aria-pressed="true"]{color:#0B1219}

/* ---------------------------------------------------------------- KPIs */
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:14px;margin:16px 0}
.kpi{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);
  padding:16px 18px;box-shadow:var(--shadow)}
.kpi .lbl{font-size:12.5px;color:var(--ink-muted)}
.kpi .v{font-size:26px;font-weight:700;letter-spacing:-.02em;margin-top:3px;
  font-variant-numeric:proportional-nums}
.kpi .v.warn{color:var(--warning)} .kpi .v.bad{color:var(--danger)}
.kpi .hint{font-size:11.5px;color:var(--ink-muted);margin-top:3px}

/* ---------------------------------------------------------------- cards */
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);
  padding:20px 22px;box-shadow:var(--shadow);min-width:0}
.card h2{font-family:var(--serif);font-size:17px;font-weight:600;margin-bottom:3px}
.card .cap{font-size:12.5px;color:var(--ink-muted);margin-bottom:16px}
/* min-width:0 nos itens: sem isso, o min-width:auto do grid impede o card de
   encolher abaixo do conteúdo e a PÁGINA rola na horizontal (o certo é a
   tabela rolar dentro do próprio container). */
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin:16px 0}
.grid2>*{min-width:0}
@media(max-width:860px){.grid2{grid-template-columns:1fr}}

/* ---------------------------------------------------------------- barras
   Série única (uma cor). Ponta de dado 4px arredondada, base reta. */
.brow{display:grid;grid-template-columns:minmax(90px,150px) 1fr 84px;gap:12px;
  align-items:center;padding:4px 6px;margin:0 -6px;border-radius:var(--r-sm);
  cursor:pointer;transition:background .12s}
.brow:hover,.brow:focus-visible{background:var(--surface-3)}
.brow:active{background:var(--track)}
.blabel{font-size:12.5px;color:var(--ink-2);white-space:nowrap;overflow:hidden;
  text-overflow:ellipsis}
.btrack{background:var(--track);border-radius:0 6px 6px 0;height:13px;overflow:hidden}
.bfill{height:100%;background:var(--primary);border-radius:0 4px 4px 0;
  transition:width .25s ease}
.brow:hover .bfill{filter:brightness(1.12)}
.bval{font-size:12px;text-align:right;color:var(--ink-2);
  font-variant-numeric:tabular-nums;white-space:nowrap}
.vazio{color:var(--ink-muted);font-size:13px;padding:14px 0;text-align:center}

/* ---------------------------------------------------------------- tabelas */
.tblwrap{overflow-x:auto;max-width:100%;-webkit-overflow-scrolling:touch}
table{width:100%;border-collapse:collapse;font-size:12.5px}
th,td{padding:8px 9px;border-bottom:1px solid var(--border);text-align:left;vertical-align:top}
thead th{position:sticky;top:0;background:var(--surface);z-index:1;
  color:var(--ink-muted);font-weight:600;white-space:nowrap;cursor:pointer;
  user-select:none;border-bottom:1.5px solid var(--border-2)}
thead th:hover{color:var(--ink)}
tbody tr:hover{background:var(--surface-2)}
#tb tbody tr,#tbPlan tbody tr{cursor:pointer}
#tb tbody tr:focus-visible,#tbPlan tbody tr:focus-visible{outline:2px solid var(--accent);outline-offset:-2px}
td.num,th.num{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
/* ---- modal de detalhes (clique na linha) ---- */
.modal-bg{position:fixed;inset:0;z-index:80;background:rgba(8,14,22,.55);
  display:flex;align-items:flex-start;justify-content:center;padding:5vh 16px;
  overflow-y:auto;-webkit-overflow-scrolling:touch}
.modal-bg[hidden]{display:none}
.modal{position:relative;background:var(--surface);color:var(--ink);
  border:1px solid var(--border-2);border-radius:14px;max-width:760px;width:100%;
  box-shadow:0 24px 60px rgba(0,0,0,.4);padding:22px 24px 26px}
.modal-x{position:absolute;top:12px;right:12px;width:34px;height:34px;border-radius:9px;
  border:1px solid var(--border-2);background:var(--surface-3);color:var(--ink-muted);
  font-size:15px;cursor:pointer;line-height:1}
.modal-x:hover{color:var(--ink);background:var(--surface)}
.modal h3{font-family:var(--sans);font-size:19px;margin:0 34px 3px 0;line-height:1.25}
.modal .m-sub{color:var(--ink-muted);font-size:13px;margin:0 0 16px}
.m-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px 22px;margin:0 0 6px}
@media(max-width:560px){.m-grid{grid-template-columns:1fr}}
.m-f{min-width:0}
.m-f .k{display:block;font-size:11px;text-transform:uppercase;letter-spacing:.03em;
  color:var(--ink-muted);font-weight:600;margin-bottom:2px}
.m-f .v{display:block;font-size:14px;color:var(--ink);overflow-wrap:anywhere}
.m-f.wide{grid-column:1/-1}
.m-sec{margin:18px 0 8px;font-size:12px;font-weight:700;text-transform:uppercase;
  letter-spacing:.04em;color:var(--ink-muted);border-top:1px solid var(--border);padding-top:14px}
.modal .tblwrap{max-height:44vh;overflow:auto}
.modal table{font-size:12px}
.m-badge{display:inline-block;padding:2px 9px;border-radius:999px;font-size:12px;font-weight:600}
.m-badge.p-ger{background:color-mix(in srgb,var(--accent) 15%,transparent);color:var(--accent)}
.m-badge.p-part{background:var(--surface-3);color:var(--ink-muted)}
.m-link{display:inline-flex;align-items:center;gap:6px;margin-top:4px;font-weight:600;color:var(--accent)}
td.desc{max-width:270px;color:var(--ink-2)}
td.forn{max-width:160px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:var(--ink-2)}
td.cat{white-space:nowrap;font-size:11.5px;color:var(--ink-muted)}
td.nd{white-space:nowrap;font-variant-numeric:tabular-nums;font-size:11.5px;color:var(--ink-muted)}
/* Status: cor + ÍCONE + rótulo (nunca cor sozinha) */
.tag{display:inline-flex;align-items:center;gap:5px;border-radius:999px;
  padding:2px 9px;font-size:11px;white-space:nowrap;font-weight:500}
.tag-vig{background:color-mix(in srgb,var(--success) 14%,transparent);color:var(--success)}
.tag-v30{background:color-mix(in srgb,var(--warning) 16%,transparent);color:var(--warning)}
.tag-venc{background:color-mix(in srgb,var(--danger) 14%,transparent);color:var(--danger)}
.tag-semata{background:var(--surface-3);color:var(--ink-muted)}
.tbl-foot{display:flex;justify-content:center;margin-top:14px}
a{color:var(--primary-ink)} a:hover{text-decoration:underline}
.spark-leg{display:flex;justify-content:space-between;font-size:11.5px;
  color:var(--ink-muted);margin-top:4px;font-variant-numeric:tabular-nums}

/* ---------------------------------------------------------------- tooltip */
.tip{position:fixed;z-index:60;pointer-events:none;opacity:0;transition:opacity .12s;
  background:var(--ink);color:var(--bg);border-radius:var(--r-sm);padding:8px 11px;
  font-size:12.5px;box-shadow:var(--shadow-2);max-width:280px}
.tip.on{opacity:1}
.tip .tv{font-weight:700;font-size:14px;display:block}
.tip .tl{color:var(--ink-muted);display:block;margin-top:1px}
[data-theme="dark"] .tip{background:#E7EDF4;color:#0B1219}
[data-theme="dark"] .tip .tl{color:#42505F}

footer{margin-top:30px;font-size:12px;color:var(--ink-muted);text-align:center;line-height:1.7}
.sr{position:absolute;width:1px;height:1px;overflow:hidden;clip:rect(0 0 0 0);white-space:nowrap}

/* ---------------------------------------------------------------- impressão */
@media print{
  .hdr,.filtros,.tbl-foot,.btn,.seg,.tip{display:none!important}
  body{background:#fff;color:#000;font-size:10pt}
  .card,.hero,.kpi{box-shadow:none;border:1px solid #999;break-inside:avoid}
  .grid2{grid-template-columns:1fr 1fr}
  .tblwrap{overflow:visible} thead th{position:static}
  a{text-decoration:none;color:#000}
}
</style>
</head>
<body>
<div class="bcms-bar" aria-hidden="true"></div>

<header class="hdr">
  <div class="hdr-in">
    <div class="hdr-l">
      <img class="emblema" id="emblema" src="%%EMBLEMA%%" alt="Brasão %%UNIDADE_CURTA%%"
           onerror="this.style.visibility='hidden'">
      <div>
        <h1>Capacidade de Empenho</h1>
        <p class="sub"><span id="uNome">%%UNIDADE%%</span> · <span id="uUasg">UASG %%UASG%%</span></p>
      </div>
    </div>
    <div class="hdr-r">
      <span class="chip-pos">Posição <b>%%POSICAO%%</b></span>
      <button class="btn" id="btTheme" aria-label="Alternar tema claro/escuro">◐ Tema</button>
    </div>
  </div>
</header>

<nav class="omds-nav" aria-label="Trocar de organização militar">
  <div class="omds-nav-in">%%OMDS_NAV%%</div>
</nav>

<main class="wrap">
  <section class="aguardando" id="semColeta" aria-live="polite">
    <img id="scEmblema" src="" alt="" onerror="this.style.display='none'">
    <h2 id="scNome"></h2>
    <p class="u-uasg" id="scUasg"></p>
    <span class="u-badge">⏳ Pregões desta OM ainda não coletados</span>
    <p class="u-hint">Assim que o robô extrator coletar as atas desta unidade,
      o painel de capacidade de empenho aparecerá aqui automaticamente.</p>
  </section>

  <div id="painel">
  <nav class="tabs" role="tablist" aria-label="Seções">
    <button class="tab" id="tabCap" role="tab" aria-selected="true"
            aria-controls="tab-capacidade">Capacidade de empenho</button>
    <button class="tab" id="tabPlan" role="tab" aria-selected="false"
            aria-controls="tab-planejamento">📅 Planejamento · atas vencendo</button>
  </nav>

  <div id="tab-capacidade" role="tabpanel">
  <section class="hero" aria-labelledby="heroLabel">
    <p class="hero-label" id="heroLabel">Capacidade de empenho disponível</p>
    <p class="hero-big" id="heroBig">%%CAP_TOTAL%%</p>
    <p class="hero-cap">Saldo das atas × valor unitário registrado — o quanto ainda
      pode ser empenhado sem nova licitação.</p>
    <div class="meter" role="img" id="meter"
         aria-label="Composição: vigentes e vencendo em até 30 dias">
      <div class="mseg mseg-vig" id="segVig" style="width:%%PCT_VIG_M%%%"></div>
      <div class="mseg mseg-v30" id="segV30" style="width:%%PCT_V30_M%%%"></div>
    </div>
    <ul class="mleg">
      <li><span class="key key-vig" aria-hidden="true"></span>Vigentes <b id="heroVig">%%CAP_VIG%%</b></li>
      <li><span class="key key-v30" aria-hidden="true"></span>Vencem em ≤30 dias
        <b id="heroV30">%%CAP_V30%%</b> <span class="muted">(<span id="heroPct">%%PCT_V30%%</span>)</span></li>
    </ul>
    <div class="pills" id="ativo"><span class="lbl">Filtros:</span><span id="ativoPills"></span></div>
  </section>

  <section class="filtros" aria-label="Filtros">
    <div class="filtros-grid">
      %%SEL_CAT%%
      %%SEL_ND%%
      %%SEL_PG%%
      %%SEL_FORN%%
      %%SEL_TP%%
    </div>
    <div class="frow2">
      <input id="busca" type="search" placeholder="Buscar item, fornecedor, pregão…"
             aria-label="Buscar no texto dos itens">
      <button class="btn" id="btLimpar">Limpar filtros</button>
    </div>
    <p class="sr" id="vivo" role="status" aria-live="polite"></p>
  </section>

  <section class="kpis">
    <div class="kpi"><div class="lbl">Itens com saldo</div>
      <div class="v" id="kItens">%%N_ITENS%%</div>
      <div class="hint">linhas de ata disponíveis</div></div>
    <div class="kpi"><div class="lbl">Pregões</div>
      <div class="v" id="kPregoes">%%N_PREGOES%%</div>
      <div class="hint">com saldo remanescente</div></div>
    <div class="kpi"><div class="lbl">Vence em ≤30 dias</div>
      <div class="v warn" id="kV30">%%CAP_V30%%</div>
      <div class="hint">usar ou perder</div></div>
    <div class="kpi"><div class="lbl">Já perdido</div>
      <div class="v bad" id="kVenc">%%CAP_VENC%%</div>
      <div class="hint">saldo em atas vencidas</div></div>
  </section>

  <div class="grid2">
    <section class="card">
      <h2>Onde está a capacidade</h2>
      <p class="cap">Por tipo de material ou serviço — 12 maiores · <b>clique para filtrar</b></p>
      <div id="barCat">%%BARRAS_CAT%%</div>
    </section>
    <section class="card">
      <h2>Por pregão</h2>
      <p class="cap">UASG gerenciadora · número — 12 maiores · <b>clique para filtrar</b></p>
      <div id="barPg">%%BARRAS_PREGAO%%</div>
    </section>
  </div>

  <div class="grid2">
    <section class="card">
      <h2>Vencendo em até 60 dias</h2>
      <p class="cap">Prioridade de empenho antes da perda do saldo</p>
      <div class="tblwrap"><table>
        <thead><tr><th>Fim vigência</th><th>Pregão</th><th class="num">Itens</th>
        <th class="num">Capacidade</th></tr></thead>
        <tbody id="tbVenc">%%VENC60%%</tbody>
      </table></div>
    </section>
    <section class="card">
      <h2>Evolução</h2>
      <p class="cap">Capacidade total a cada atualização</p>
      %%SPARK%%
    </section>
  </div>

  <section class="card">
    <h2>Itens <span class="cap" id="cont" style="font-family:var(--sans)"></span></h2>
    <p class="cap"><b>Clique numa linha</b> para ver todos os detalhes do item ·
      clique nos títulos para ordenar · ↗ abre o item no Compras.gov.br. Atas já
      vencidas não são listadas.</p>
    <div class="frow2" style="margin:0 0 14px;align-items:center">
      <div class="seg" id="segSit" role="group" aria-label="Situação da ata (filtra apenas esta lista)">
        <button data-f="ativos" aria-pressed="true">Válidas</button>
        <button data-f="v30" aria-pressed="false">≤30 dias</button>
        <button data-f="all" aria-pressed="false">Todas</button>
      </div>
      <button class="btn btn-pregao" id="btPregao" aria-pressed="false" hidden></button>
      <span class="cap" style="margin:0">Vigentes e vencendo — sem as vencidas.</span>
      <button type="button" class="btn" id="btCsvItens" style="margin-left:auto"
              title="Baixar em Excel (.xlsx) a tabela com o filtro atual">⬇ Exportar Excel</button>
    </div>
    <div class="tblwrap">
      <table id="tb">
        <thead><tr>
          <th>Pregão</th><th>Ger.</th><th class="num">Item</th><th>Descrição</th>
          <th>Categoria</th><th>ND sug.</th><th>Fornecedor</th><th>Fim vig.</th>
          <th>Situação</th><th class="num">Saldo</th><th class="num">Vlr. unit.</th>
          <th class="num">Valor total</th><th><span class="sr">Link</span></th>
        </tr></thead>
        <tbody>
%%TABELA%%
        </tbody>
      </table>
    </div>
    <p class="vazio" id="semResultado" style="display:none">
      Nenhum item para esta combinação de filtros.</p>
    <div class="tbl-foot"><button class="btn" id="btMais" style="display:none"></button></div>
  </section>
  </div><!-- /tab-capacidade -->

  <div id="tab-planejamento" role="tabpanel" hidden>
    <section class="hero" style="--accent:var(--warning-fill)">
      <p class="hero-label">Planejamento de novas licitações</p>
      <p class="hero-big" id="planTitulo" style="font-size:clamp(26px,4vw,34px)">—</p>
      <p class="hero-cap">Atas vencendo, agrupadas por pregão e ordenadas pelo
        vencimento mais próximo. Use para o setor se antecipar e abrir novos
        pregões antes que o registro de preços expire.</p>
    </section>

    <section class="card">
      <div class="pl-bandas" id="planBandas" role="group"
           aria-label="Resumo por prazo — clique para filtrar a lista">
        <button type="button" class="pl-card b-30" data-band="b30" aria-pressed="false">
          <span class="pl-card-n" data-n>0</span>
          <span class="pl-card-lb">vencendo em ≤30 dias</span>
          <span class="pl-card-cap" data-cap>—</span></button>
        <button type="button" class="pl-card b-90" data-band="b90" aria-pressed="false">
          <span class="pl-card-n" data-n>0</span>
          <span class="pl-card-lb">vencem em 31–90 dias</span>
          <span class="pl-card-cap" data-cap>—</span></button>
        <button type="button" class="pl-card b-180" data-band="b180" aria-pressed="false">
          <span class="pl-card-n" data-n>0</span>
          <span class="pl-card-lb">vencem em 91–180 dias</span>
          <span class="pl-card-cap" data-cap>—</span></button>
        <button type="button" class="pl-card b-max" data-band="bmax" aria-pressed="false">
          <span class="pl-card-n" data-n>0</span>
          <span class="pl-card-lb">vigentes &gt; 180 dias</span>
          <span class="pl-card-cap" data-cap>—</span></button>
      </div>
      <div class="frow2" style="margin:0 0 8px;align-items:center">
        <div class="seg" id="planPapel" role="group" aria-label="Papel do BCMS no pregão">
          <button data-p="all" aria-pressed="true">Todos</button>
          <button data-p="ger" aria-pressed="false">Gerenciador</button>
          <button data-p="part" aria-pressed="false">Participante</button>
        </div>
        <span class="cap" style="margin:0" id="planResumo"></span>
        <button type="button" class="btn" id="btPlanCsv" style="margin-left:auto"
                title="Baixar a lista visível em Excel (.xlsx)">⬇ Exportar Excel</button>
      </div>
      <p class="cap">Cartões acima resumem os pregões por prazo — <b>clique</b> em um
        para filtrar a lista (clique de novo para ver todos). <b>Gerencia</b> = o BCMS
        conduz o pregão (precisa abrir a nova licitação); <b>participa</b> = carona de
        pregão de outra OM. O "objeto" é resumido pelas categorias dos itens (o
        Compras.gov não traz o objeto do pregão). Clique nos títulos para ordenar.</p>
      <div class="tblwrap">
        <table id="tbPlan">
          <thead><tr>
            <th>Vencimento</th><th>Pregão</th><th>Ger.</th><th>Objeto (categorias)</th>
            <th class="num">Itens</th><th class="num">Capacidade</th><th>Fornecedor</th>
          </tr></thead>
          <tbody>
%%TABELA_PLANEJ%%
          </tbody>
        </table>
      </div>
      <p class="vazio" id="planVazio" style="display:none">
        Nenhuma ata vencendo neste horizonte.</p>
    </section>
  </div><!-- /tab-planejamento -->

  </div><!-- /painel -->

  <footer>
    Dados extraídos do Compras.gov.br pelo Robô Extrator de Pregões.<br>
    Gerado em %%GERADO%% · Capacidade = Qtd. Saldo × Valor Unitário das atas registradas.<br>
    A Natureza de Despesa é <b>sugerida automaticamente</b> pela descrição do item — confira antes de empenhar.
  </footer>
</main>

<div class="tip" id="tip" role="tooltip" aria-hidden="true"></div>

<div class="modal-bg" id="modal" hidden>
  <div class="modal" role="dialog" aria-modal="true" aria-labelledby="modalTit">
    <button class="modal-x" id="modalX" aria-label="Fechar detalhes">✕</button>
    <div id="modalBody"></div>
  </div>
</div>

<script>
(function(){
  "use strict";
  var root=document.documentElement, KEY='bcms-pregoes-theme';
  try{var t=localStorage.getItem(KEY);
      if(t) root.dataset.theme=t;
      else if(matchMedia('(prefers-color-scheme: dark)').matches) root.dataset.theme='dark';}catch(e){}
  document.getElementById('btTheme').onclick=function(){
    root.dataset.theme = root.dataset.theme==='dark'?'light':'dark';
    try{localStorage.setItem(KEY, root.dataset.theme);}catch(e){}
  };

  // Troca de OMDS: cada unidade é uma PÁGINA (index.html / <slug>.html), já
  // pronta no servidor (tema, brasão, dados). Os brasões são links <a> — a
  // navegação normal do navegador cuida da troca; nada a fazer aqui via JS.

  var LOTE=120;                       // linhas renderizadas por vez
  var tb=document.getElementById('tb'), corpo=tb.tBodies[0];
  var dados=[].slice.call(corpo.rows).map(function(tr){
    return {tr:tr, st:tr.dataset.st, cat:tr.dataset.cat, nd:tr.dataset.nd,
            pg:tr.dataset.pg, tipo:tr.dataset.tipo, forn:tr.dataset.forn||'',
            key:tr.dataset.key,
            cap:parseFloat(tr.dataset.cap)||0, saldo:parseFloat(tr.dataset.saldo)||0,
            fim:parseFloat(tr.dataset.fimts)||0, txt:tr.textContent.toLowerCase()};
  });
  var fCat=document.getElementById('fCat'), fNd=document.getElementById('fNd'),
      fPg=document.getElementById('fPg'), fTp=document.getElementById('fTp'),
      fForn=document.getElementById('fForn'),
      busca=document.getElementById('busca'), btMais=document.getElementById('btMais'),
      btPregao=document.getElementById('btPregao'),
      status='ativos', verTudo=false, verTudoKey='', mostrando=LOTE, filtrados=[];

  function brl(v){return v.toLocaleString('pt-BR',{style:'currency',currency:'BRL'});}

  // ===== gerador de .xlsx (Excel real) em JS puro — o xlsx é um ZIP de XMLs =====
  function _crc32(buf){
    var t=_crc32.t; if(!t){t=_crc32.t=[];for(var n=0;n<256;n++){var c=n;for(var k=0;k<8;k++)c=c&1?(0xEDB88320^(c>>>1)):(c>>>1);t[n]=c>>>0;}}
    var crc=-1; for(var i=0;i<buf.length;i++) crc=(crc>>>8)^t[(crc^buf[i])&255]; return (crc^-1)>>>0;
  }
  function _u8(s){return new TextEncoder().encode(s);}
  function _zip(files){                      // STORE (sem compressão); Excel lê ok
    var parts=[],cen=[],off=0;
    function u16(n){return [n&255,(n>>8)&255];}
    function u32(n){return [n&255,(n>>8)&255,(n>>16)&255,(n>>24)&255];}
    files.forEach(function(f){
      var nm=_u8(f.name),crc=_crc32(f.data),sz=f.data.length,lho=off;
      var lh=new Uint8Array([].concat(u32(0x04034b50),u16(20),u16(0),u16(0),u16(0),u16(0),u32(crc),u32(sz),u32(sz),u16(nm.length),u16(0)));
      parts.push(lh,nm,f.data); off+=lh.length+nm.length+sz;
      cen.push(new Uint8Array([].concat(u32(0x02014b50),u16(20),u16(20),u16(0),u16(0),u16(0),u16(0),u32(crc),u32(sz),u32(sz),u16(nm.length),u16(0),u16(0),u16(0),u16(0),u32(0),u32(lho))),nm);
    });
    var cs=0; cen.forEach(function(c){cs+=c.length;});
    var eo=new Uint8Array([].concat(u32(0x06054b50),u16(0),u16(0),u16(files.length),u16(files.length),u32(cs),u32(off),u16(0)));
    var all=parts.concat(cen,[eo]),tot=0; all.forEach(function(a){tot+=a.length;});
    var out=new Uint8Array(tot),p=0; all.forEach(function(a){out.set(a,p);p+=a.length;}); return out;
  }
  function _xe(s){return (''+s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
  function baixarExcel(nome, cab, linhas, colsNum){
    colsNum=colsNum||[];
    function col(i){return String.fromCharCode(65+i);}      // até 26 colunas (A..Z)
    var todas=[cab].concat(linhas), rows='';
    todas.forEach(function(r,ri){
      var cs='';
      r.forEach(function(v,ci){
        if(v===''||v==null) return;
        var ref=col(ci)+(ri+1);
        if(ri>0 && colsNum.indexOf(ci)>=0 && !isNaN(parseFloat(v)))
          cs+='<c r="'+ref+'"><v>'+parseFloat(v)+'</v></c>';
        else
          cs+='<c r="'+ref+'" t="inlineStr"><is><t xml:space="preserve">'+_xe(v)+'</t></is></c>';
      });
      rows+='<row r="'+(ri+1)+'">'+cs+'</row>';
    });
    var sheet='<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>'+rows+'</sheetData></worksheet>';
    var ct='<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>';
    var rels='<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>';
    var wb='<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="Dados" sheetId="1" r:id="rId1"/></sheets></workbook>';
    var wbr='<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/></Relationships>';
    var zip=_zip([
      {name:'[Content_Types].xml',data:_u8(ct)},
      {name:'_rels/.rels',data:_u8(rels)},
      {name:'xl/workbook.xml',data:_u8(wb)},
      {name:'xl/_rels/workbook.xml.rels',data:_u8(wbr)},
      {name:'xl/worksheets/sheet1.xml',data:_u8(sheet)}
    ]);
    var blob=new Blob([zip],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
    var url=URL.createObjectURL(blob),a=document.createElement('a');
    a.href=url; a.download=nome+'.xlsx';
    document.body.appendChild(a); a.click(); document.body.removeChild(a);
    setTimeout(function(){URL.revokeObjectURL(url);},1500);
  }
  function curto(v){
    if(Math.abs(v)>=1e6) return 'R$ '+(v/1e6).toFixed(2).replace('.',',')+' mi';
    if(Math.abs(v)>=1e3) return 'R$ '+Math.round(v/1e3)+' mil';
    return brl(v);
  }
  function txt(id,v){document.getElementById(id).textContent=v;}
  function elem(tag,cls,texto){
    var e=document.createElement(tag);
    if(cls)e.className=cls;
    if(texto!=null)e.textContent=texto;   // rótulos vêm da planilha: nunca innerHTML
    return e;
  }

  // ---- tooltip das barras (o valor também está rotulado na ponta) ----
  var tip=document.getElementById('tip');
  function mostraTip(ev,lab,val){
    tip.textContent='';
    tip.appendChild(elem('span','tv',val));
    tip.appendChild(elem('span','tl',lab));
    tip.classList.add('on');
    var r=(ev.currentTarget||ev.target).getBoundingClientRect();
    var x=(ev.clientX||r.left+r.width/2), y=r.top;
    tip.style.left=Math.min(Math.max(10,x-tip.offsetWidth/2),innerWidth-tip.offsetWidth-10)+'px';
    tip.style.top=Math.max(8,y-tip.offsetHeight-10)+'px';
  }
  function escondeTip(){tip.classList.remove('on');}

  function barras(el,mapa,tipoFiltro){
    var arr=Object.keys(mapa).map(function(k){return [k,mapa[k].cap,mapa[k].val];})
              .sort(function(a,b){return b[1]-a[1];}).slice(0,12);
    el.textContent='';
    if(!arr.length){el.appendChild(elem('p','vazio','Sem dados para este filtro.'));return;}
    var topo=arr[0][1]||1;
    arr.forEach(function(p){
      var row=elem('div','brow'); row.tabIndex=0; row.setAttribute('role','button');
      row.setAttribute('aria-label','Filtrar por '+p[0]); row.title='Filtrar por '+p[0];
      row.appendChild(elem('div','blabel',p[0]));
      var tr=elem('div','btrack'), fl=elem('div','bfill');
      fl.style.width=Math.max(p[1]/topo*100,1.5).toFixed(1)+'%';
      tr.appendChild(fl); row.appendChild(tr);
      row.appendChild(elem('div','bval',curto(p[1])));
      var lab=p[0], val=brl(p[1]), fval=p[2];
      row.addEventListener('pointermove',function(e){mostraTip(e,lab,val);});
      row.addEventListener('pointerleave',escondeTip);
      row.addEventListener('focus',function(e){mostraTip(e,lab,val);});
      row.addEventListener('blur',escondeTip);
      // clicar na barra aplica o filtro correspondente na lista abaixo
      row.addEventListener('click',function(){escondeTip(); aplicaFiltro(tipoFiltro,fval);});
      row.addEventListener('keydown',function(e){
        if(e.key==='Enter'||e.key===' '){e.preventDefault(); escondeTip(); aplicaFiltro(tipoFiltro,fval);}});
      el.appendChild(row);
    });
  }

  // Clique numa barra → aplica o filtro de categoria ou de pregão e rola até a
  // lista, para o usuário ver os itens daquele grupo.
  function aplicaFiltro(tipo, val){
    var sel = (tipo==='cat') ? fCat : (tipo==='pg') ? fPg : null;
    if(sel){ sel.value=val; if(sel._refletirCombo) sel._refletirCombo(); }
    aplica();
    var tbl=document.getElementById('tb');
    var card=tbl.closest ? tbl.closest('.card') : null;
    (card||tbl).scrollIntoView({behavior:'smooth', block:'start'});
  }

  // ---- combobox pesquisável: transforma um <select> num campo com busca ----
  function montarCombo(sel){
    var wrap=document.createElement('div'); wrap.className='cb';
    var inp=document.createElement('input'); inp.type='text'; inp.className='cb-in';
    inp.setAttribute('role','combobox'); inp.setAttribute('aria-autocomplete','list');
    inp.setAttribute('aria-expanded','false'); inp.autocomplete='off';
    var lista=document.createElement('ul'); lista.className='cb-list'; lista.hidden=true;
    lista.setAttribute('role','listbox');
    var clr=document.createElement('button'); clr.type='button'; clr.className='cb-clear';
    clr.setAttribute('aria-label','Limpar'); clr.textContent='×'; clr.hidden=true;
    var opts=[].slice.call(sel.options).map(function(o){return {val:o.value, txt:o.text};});
    var placeholder=opts.length?opts[0].txt:'';
    inp.placeholder=placeholder;

    function textoDe(val){
      for(var i=0;i<opts.length;i++) if(opts[i].val===val) return opts[i].txt;
      return '';
    }
    function refletir(){                       // sincroniza input ↔ select
      inp.value = sel.value==='' ? '' : textoDe(sel.value);
      clr.hidden = sel.value==='';
    }
    function render(filtro){
      lista.textContent='';
      var f=(filtro||'').toLowerCase();
      var achou=0;
      opts.forEach(function(o){
        // busca no texto E no valor (o valor do fornecedor tem CNPJ+nome completos)
        if(f && (o.txt+' '+o.val).toLowerCase().indexOf(f)===-1) return;
        achou++;
        var li=document.createElement('li'); li.className='cb-opt'; li.setAttribute('role','option');
        li.textContent=o.txt; li.dataset.val=o.val;
        if(o.val===sel.value) li.setAttribute('aria-selected','true');
        li.addEventListener('mousedown',function(e){e.preventDefault(); escolher(o.val);});
        lista.appendChild(li);
      });
      if(!achou){var li=document.createElement('li'); li.className='cb-vazio';
                 li.textContent='Nada encontrado'; lista.appendChild(li);}
    }
    function abrir(){ render(''); lista.hidden=false; inp.setAttribute('aria-expanded','true'); }
    function fechar(){ lista.hidden=true; inp.setAttribute('aria-expanded','false'); refletir(); }
    function escolher(val){ sel.value=val;
      sel.dispatchEvent(new Event('change')); refletir(); fechar(); }

    inp.addEventListener('focus',function(){ inp.select(); abrir(); });
    inp.addEventListener('input',function(){ lista.hidden=false; clr.hidden=inp.value===''; render(inp.value); });
    inp.addEventListener('keydown',function(e){
      if(e.key==='Escape'){ fechar(); inp.blur(); }
      else if(e.key==='Enter'){ var o=lista.querySelector('.cb-opt');
        if(o){ e.preventDefault(); escolher(o.dataset.val); } }
    });
    inp.addEventListener('blur',function(){ setTimeout(fechar,140); }); // deixa o mousedown rodar
    clr.addEventListener('click',function(){ escolher(''); inp.focus(); });

    sel._refletirCombo=refletir;               // p/ cliques nas barras e "Limpar"
    sel.style.display='none';
    sel.parentNode.insertBefore(wrap, sel);
    wrap.appendChild(inp); wrap.appendChild(clr); wrap.appendChild(lista); wrap.appendChild(sel);
    refletir();
  }

  function pinta(){
    filtrados.forEach(function(d,i){d.tr.style.display = i<mostrando ? '' : 'none';});
    var resta=filtrados.length-mostrando;
    btMais.style.display = resta>0 ? '' : 'none';
    if(resta>0) btMais.textContent='Mostrar mais '+Math.min(resta,LOTE)+
      ' (de '+resta.toLocaleString('pt-BR')+' restantes)';
    document.getElementById('semResultado').style.display=filtrados.length?'none':'';
  }

  function aplica(reset){
    if(reset!==false) mostrando=LOTE;
    var vc=fCat.value, vn=fNd.value, vp=fPg.value, vt=fTp.value, vf=fForn.value,
        termo=busca.value.trim().toLowerCase();
    // "Ver pregão completo" só vale para o pregão em que foi ligado; ao trocar
    // ou limpar o pregão, volta à visão de capacidade.
    if(vp!==verTudoKey) verTudo=false;
    var capV=0, capW=0, capVenc=0, cats={}, pgs={}, chaves={}, nItens=0;
    var lim=Date.now()/1000+60*86400, venc={};
    filtrados=[];

    dados.forEach(function(d){
      // DIMENSÃO (categoria/ND/pregão/empresa/tipo/busca) governa TODOS os
      // números do resumo — inclusive "já perdido", que olha as vencidas.
      // O pregão é casado pela CHAVE (ger · nº) — pregão preciso, com a UASG.
      var okDim = (!vc||d.cat===vc) && (!vn||d.nd===vn) && (!vp||d.key===vp)
                && (!vt||d.tipo===vt) && (!vf||d.forn===vf) && (!termo||d.txt.indexOf(termo)>-1);
      // "Item de capacidade" = tem saldo E não é "sem dados" (era o que a lista
      // mostrava antes). SITUAÇÃO é só um seletor de visualização da lista.
      var temItem = d.saldo>0 && d.st!=='semdados';
      // Itens VENCIDOS nunca entram na lista (o usuário não quer vê-los); seguem
      // no DOM só para alimentar o KPI "Já perdido".
      var okS = d.st!=='venc' && ((status==='all') || (status==='ativos' && (d.st==='vig'||d.st==='v30'))
              || (status==='v30' && d.st==='v30'));
      // No modo "todos do pregão", a lista mostra TODO item do pregão (menos as
      // vencidas); na visão normal, só itens com saldo e da situação.
      var naLista = verTudo ? (okDim && d.st!=='venc') : (okDim && temItem && okS);
      d.tr.style.display = naLista ? '' : 'none';
      if(naLista) filtrados.push(d);
      // Os NÚMEROS do resumo só contam itens com capacidade (saldo × vu > 0),
      // independentemente do modo de exibição da lista.
      if(!okDim || !(d.cap>0)) return;
      if(d.st==='venc'){capVenc+=d.cap;return;}
      if(d.st==='vig'){capV+=d.cap;} else if(d.st==='v30'){capW+=d.cap;} else {return;}
      nItens++; chaves[d.key]=1;
      // guarda {cap, val}: val é o que o filtro recebe ao clicar na barra
      (cats[d.cat]=cats[d.cat]||{cap:0,val:d.cat}).cap+=d.cap;
      (pgs[d.key]=pgs[d.key]||{cap:0,val:d.key}).cap+=d.cap;
      if(d.fim && d.fim<lim){
        var k=d.fim+'|'+d.key, v=venc[k]||(venc[k]={cap:0,n:0,fim:d.fim,key:d.key});
        v.cap+=d.cap; v.n++;
      }
    });
    pinta();

    var total=capV+capW;
    txt('heroBig',brl(total)); txt('heroVig',brl(capV)); txt('heroV30',brl(capW));
    txt('heroPct',(total?(capW/total*100).toFixed(1).replace('.',','):'0,0')+'%');
    document.getElementById('segVig').style.width=(total?capV/total*100:0).toFixed(2)+'%';
    document.getElementById('segV30').style.width=(total?capW/total*100:0).toFixed(2)+'%';
    txt('kItens',nItens.toLocaleString('pt-BR'));
    txt('kPregoes',Object.keys(chaves).length.toLocaleString('pt-BR'));
    txt('kV30',brl(capW)); txt('kVenc',brl(capVenc));
    txt('cont','· '+filtrados.length.toLocaleString('pt-BR')+' listados');
    txt('vivo',filtrados.length.toLocaleString('pt-BR')+' itens; capacidade '+brl(total));

    barras(document.getElementById('barCat'),cats,'cat');
    barras(document.getElementById('barPg'),pgs,'pg');

    var alvo=document.getElementById('tbVenc'); alvo.textContent='';
    var vs=Object.keys(venc).map(function(k){return venc[k];})
            .sort(function(a,b){return a.fim-b.fim;}).slice(0,12);
    if(!vs.length){
      var tr0=document.createElement('tr'), td0=elem('td','vazio',
        'Nenhuma ata vencendo nos próximos 60 dias.');
      td0.colSpan=4; tr0.appendChild(td0); alvo.appendChild(tr0);
    } else vs.forEach(function(v){
      var d=new Date(v.fim*1000), tr=document.createElement('tr');
      tr.appendChild(elem('td',null,('0'+d.getDate()).slice(-2)+'/'+
        ('0'+(d.getMonth()+1)).slice(-2)+'/'+d.getFullYear()));
      tr.appendChild(elem('td',null,v.key));
      tr.appendChild(elem('td','num',String(v.n)));
      var tdc=elem('td','num'); tdc.appendChild(elem('strong',null,brl(v.cap)));
      tr.appendChild(tdc); alvo.appendChild(tr);
    });

    var box=document.getElementById('ativo'), alvoP=document.getElementById('ativoPills');
    alvoP.textContent='';
    var pills=[];
    if(vc)pills.push(vc); if(vn)pills.push('ND '+vn);
    if(vp)pills.push('Pregão '+vp); if(vt)pills.push(vt);
    if(vf){var nomeF=vf.indexOf(' - ')>-1?vf.split(' - ').slice(1).join(' - '):vf;
           pills.push('Empresa: '+(nomeF.length>34?nomeF.slice(0,33)+'…':nomeF));}
    if(termo)pills.push('“'+busca.value.trim()+'”');
    pills.forEach(function(p){alvoP.appendChild(elem('span','pill',p));});
    box.classList.toggle('on',pills.length>0);
    txt('heroLabel',pills.length?'Capacidade de empenho (filtrada)'
                                :'Capacidade de empenho disponível');

    // Botão "todos os itens do pregão": só quando 1 pregão está selecionado.
    if(vp){
      btPregao.hidden=false;
      btPregao.setAttribute('aria-pressed', verTudo?'true':'false');
      btPregao.textContent = verTudo
        ? '✓ Mostrando o pregão completo — voltar aos itens com saldo'
        : 'Mostrar todos os itens do pregão';
      // Em modo "pregão completo", os chips de situação não se aplicam.
      document.getElementById('segSit').style.opacity = verTudo ? '.45' : '';
      document.getElementById('segSit').style.pointerEvents = verTudo ? 'none' : '';
    } else {
      btPregao.hidden=true;
      document.getElementById('segSit').style.opacity='';
      document.getElementById('segSit').style.pointerEvents='';
    }
  }

  [fCat,fNd,fPg,fForn,fTp].forEach(function(s){s.addEventListener('change',function(){aplica();});});
  busca.addEventListener('input',function(){aplica();});
  // Torna os selects longos pesquisáveis (digite para achar). Aprimoramento
  // progressivo: se o JS falhar, o <select> puro continua funcionando.
  [fCat,fNd,fPg,fForn].forEach(montarCombo);
  [].forEach.call(document.querySelectorAll('#segSit button'),function(b){
    b.onclick=function(){
      [].forEach.call(document.querySelectorAll('#segSit button'),function(x){
        x.setAttribute('aria-pressed','false');});
      b.setAttribute('aria-pressed','true'); status=b.dataset.f;
      verTudo=false; aplica();};});   // escolher situação sai do "pregão completo"
  btPregao.onclick=function(){
    verTudo=!verTudo; verTudoKey=fPg.value; aplica();
    document.getElementById('tb').scrollIntoView({behavior:'smooth', block:'start'});
  };
  document.getElementById('btLimpar').onclick=function(){
    fCat.value=fNd.value=fPg.value=fForn.value=fTp.value=''; busca.value='';
    [fCat,fNd,fPg,fForn].forEach(function(s){if(s._refletirCombo)s._refletirCombo();});
    [].forEach.call(document.querySelectorAll('#segSit button'),function(x){
      x.setAttribute('aria-pressed', x.dataset.f==='ativos'?'true':'false');});
    status='ativos'; aplica();};
  btMais.onclick=function(){mostrando+=LOTE; pinta();};

  // Exporta a tabela do FILTRO ATUAL (todos os itens de `filtrados`, não só a
  // página visível) em EXCEL (.xlsx). Colunas completas lidas dos data-* das
  // linhas (mesmos campos do detalhe); valores/quantidades como número.
  document.getElementById('btCsvItens').onclick=function(){
    if(!filtrados.length){ alert('Não há itens no filtro atual para exportar.'); return; }
    var cab=['Pregao','UASG_Gerenciadora','Item','Descricao','CATMAT_Nome','CATMAT_Codigo',
             'Categoria','ND_Sugerida','Subitem','Confianca_ND','Fornecedor','Nr_Ata',
             'Inicio_Vigencia','Fim_Vigencia','Situacao','Qtd_Homologada','Qtd_Autorizada',
             'Saldo','Valor_Unitario','Valor_Total_Homologado','Capacidade_saldoxunit','Link'];
    var linhas=filtrados.map(function(d){
      var tr=d.tr, x=tr.dataset;
      return [tr.cells[0].textContent.trim(), tr.cells[1].textContent.trim(),
              tr.cells[2].textContent.trim(), x.desc||'', x.catmat||'', x.cod||'',
              x.cat||'', x.nd||'', x.sub||'', x.conf||'', x.forn||'', x.nrata||'',
              x.ini||'', x.fim||'', x.stlabel||'', x.qhom||'', x.qaut||'', x.saldo||'',
              x.vu||'', x.vtot||'', x.cap||'', x.link||''];
    });
    var hoje=new Date().toISOString().slice(0,10);
    baixarExcel('capacidade_itens_%%UNIDADE_CURTA%%_'+hoje, cab, linhas, [15,16,17,18,19,20]);
  };

  var ord={col:-1,asc:false};
  [].forEach.call(tb.tHead.rows[0].cells,function(th,i){
    th.onclick=function(){
      ord.asc = ord.col===i ? !ord.asc : false; ord.col=i;
      var isNum = th.classList.contains('num') || i===7;
      dados.sort(function(a,b){
        var ca=a.tr.cells[i], cb=b.tr.cells[i], va, vb;
        if(isNum){va=parseFloat(ca.dataset.v||0)||0; vb=parseFloat(cb.dataset.v||0)||0;}
        else {va=ca.textContent.toLowerCase(); vb=cb.textContent.toLowerCase();}
        return (va<vb?-1:va>vb?1:0)*(ord.asc?1:-1);
      });
      var frag=document.createDocumentFragment();
      dados.forEach(function(d){frag.appendChild(d.tr);});
      corpo.appendChild(frag);
      aplica(false);
    };});

  aplica();

  // ==================== ABAS + PLANEJAMENTO ====================
  var tabCap=document.getElementById('tabCap'), tabPlan=document.getElementById('tabPlan'),
      painelCap=document.getElementById('tab-capacidade'),
      painelPlan=document.getElementById('tab-planejamento');
  function trocaAba(ehPlan){
    tabPlan.setAttribute('aria-selected', ehPlan?'true':'false');
    tabCap.setAttribute('aria-selected', ehPlan?'false':'true');
    painelPlan.hidden=!ehPlan; painelCap.hidden=ehPlan;
    if(ehPlan) aplicaPlanej();
    window.scrollTo(0,0);
    try{localStorage.setItem('bcms-aba', ehPlan?'plan':'cap');}catch(e){}
  }
  tabCap.onclick=function(){trocaAba(false);};
  tabPlan.onclick=function(){trocaAba(true);};

  var planTb=document.getElementById('tbPlan'), planCorpo=planTb.tBodies[0];
  var planLinhas=[].slice.call(planCorpo.rows), planP='all', planBand=null;
  function diasDe(ts){ return Math.ceil((ts*1000 - Date.now())/86400000); }
  function bandaDe(d){ return d<=30?'b30':d<=90?'b90':d<=180?'b180':'bmax'; }
  var BANDA_LB={b30:'vencendo em ≤30 dias', b90:'vencendo em 31–90 dias',
                b180:'vencendo em 91–180 dias', bmax:'vigentes (mais de 180 dias)'};
  function aplicaPlanej(){
    var tot={b30:{n:0,c:0},b90:{n:0,c:0},b180:{n:0,c:0},bmax:{n:0,c:0}};
    var n=0, cap=0, nGer=0, nPart=0;
    planLinhas.forEach(function(tr){
      var ts=parseFloat(tr.dataset.fimts)||0, dias=diasDe(ts);
      var span=tr.querySelector('.pl-dias');
      if(span){
        if(dias<0){span.textContent='venceu há '+(-dias)+' d'; span.className='pl-dias d-venc';}
        else if(dias===0){span.textContent='vence hoje'; span.className='pl-dias d-venc';}
        else if(dias<=30){span.textContent='em '+dias+' dias'; span.className='pl-dias d-30';}
        else {span.textContent='em '+dias+' dias'; span.className='pl-dias d-ok';}
      }
      var band=bandaDe(dias), capv=parseFloat(tr.cells[5].dataset.v)||0;
      var okP = planP==='all' || tr.dataset.papel===planP;
      if(okP){ tot[band].n++; tot[band].c+=capv; }   // cartões seguem o papel
      var ok = okP && (planBand===null || band===planBand);
      tr.style.display = ok?'':'none';
      if(ok){ n++; cap+=capv; if(tr.dataset.papel==='ger') nGer++; else nPart++; }
    });
    [].forEach.call(document.querySelectorAll('#planBandas .pl-card'),function(c){
      var b=c.dataset.band;
      c.querySelector('[data-n]').textContent=tot[b].n.toLocaleString('pt-BR');
      c.querySelector('[data-cap]').textContent=tot[b].c>0?brl(tot[b].c):'sem saldo';
    });
    document.getElementById('planBandas').classList.toggle('tem-sel', planBand!==null);
    document.getElementById('planVazio').style.display = n?'none':'';
    var papelTxt = planP==='ger'?' · só gerenciados pelo BCMS'
                 : planP==='part'?' · só participações (carona)'
                 : (' · '+nGer+' geren. + '+nPart+' partic.');
    document.getElementById('planResumo').textContent =
      n.toLocaleString('pt-BR')+' pregões · '+brl(cap)+' em capacidade'+papelTxt;
    document.getElementById('planTitulo').textContent = planBand===null
      ? n.toLocaleString('pt-BR')+' pregões com atas a renovar'
      : n.toLocaleString('pt-BR')+' pregões '+BANDA_LB[planBand];
  }
  [].forEach.call(document.querySelectorAll('#planBandas .pl-card'),function(b){
    b.onclick=function(){
      var jaSel=b.getAttribute('aria-pressed')==='true';
      [].forEach.call(document.querySelectorAll('#planBandas .pl-card'),function(x){
        x.setAttribute('aria-pressed','false');});
      if(jaSel){ planBand=null; }
      else { b.setAttribute('aria-pressed','true'); planBand=b.dataset.band; }
      aplicaPlanej();};});
  [].forEach.call(document.querySelectorAll('#planPapel button'),function(b){
    b.onclick=function(){
      [].forEach.call(document.querySelectorAll('#planPapel button'),function(x){
        x.setAttribute('aria-pressed','false');});
      b.setAttribute('aria-pressed','true'); planP=b.dataset.p; aplicaPlanej();};});
  // Exporta a lista VISÍVEL (respeita banda + papel) em EXCEL (.xlsx). Usa os
  // dados completos das linhas (não o texto cortado); Itens/Capacidade número.
  document.getElementById('btPlanCsv').onclick=function(){
    var cab=['Vencimento','Situacao','Pregao','UASG_Gerenciadora','Papel','Objeto',
             'Itens','Capacidade_saldo_R$','Fornecedores'];
    var linhas=[];
    planLinhas.forEach(function(tr){
      if(tr.style.display==='none') return;
      var venc=(tr.cells[0].querySelector('b')||{}).textContent||'';
      var sit=(tr.querySelector('.pl-dias')||{}).textContent||'';
      var pg=tr.cells[1].textContent.trim();
      var ger=tr.cells[2].dataset.ger||(tr.cells[2].querySelector('b')||tr.cells[2]).textContent.trim();
      var papel=tr.dataset.papel==='ger'?'Gerenciador':'Participante';
      var obj=(tr.cells[3].getAttribute('title')||tr.cells[3].textContent).trim();
      var itens=tr.cells[4].dataset.v||tr.cells[4].textContent.trim();
      var capv=tr.cells[5].dataset.v||'0';
      var forn=(tr.cells[6].getAttribute('title')||tr.cells[6].textContent).trim();
      linhas.push([venc,sit,pg,ger,papel,obj,itens,capv,forn]);
    });
    if(!linhas.length){ alert('Nada para exportar neste filtro.'); return; }
    var hoje=new Date().toISOString().slice(0,10);
    baixarExcel('planejamento_atas_%%UNIDADE_CURTA%%_'+hoje, cab, linhas, [6,7]);
  };
  var pord={col:-1,asc:false};
  [].forEach.call(planTb.tHead.rows[0].cells,function(th,i){
    th.style.cursor='pointer';
    th.onclick=function(){
      pord.asc = pord.col===i ? !pord.asc : false; pord.col=i;
      var isNum = th.classList.contains('num');
      planLinhas.sort(function(a,b){
        var va,vb;
        if(i===0){va=parseFloat(a.dataset.fimts)||0; vb=parseFloat(b.dataset.fimts)||0;}
        else if(isNum){va=parseFloat(a.cells[i].dataset.v||0)||0; vb=parseFloat(b.cells[i].dataset.v||0)||0;}
        else {va=a.cells[i].textContent.toLowerCase(); vb=b.cells[i].textContent.toLowerCase();}
        return (va<vb?-1:va>vb?1:0)*(pord.asc?1:-1);
      });
      var frag=document.createDocumentFragment();
      planLinhas.forEach(function(tr){frag.appendChild(tr);});
      planCorpo.appendChild(frag);
    };});
  aplicaPlanej();                         // calcula os números já no carregamento
  try{ if(localStorage.getItem('bcms-aba')==='plan') trocaAba(true); }catch(e){}

  // ==================== DETALHAMENTO (modal ao clicar) ====================
  var OMNOMES=%%OMNOMES%%;
  var modal=document.getElementById('modal'), modalBody=document.getElementById('modalBody'),
      ultimoFoco=null;
  function mk(tag,cls,txt){var e=document.createElement(tag); if(cls)e.className=cls;
    if(txt!=null)e.textContent=txt; return e;}
  function campo(k,v,wide){var f=mk('div','m-f'+(wide?' wide':''));
    f.appendChild(mk('span','k',k));
    f.appendChild(mk('span','v',(v==null||v==='')?'—':v)); return f;}
  function money(x){var n=parseFloat(x); return isFinite(n)?brl(n):'—';}
  function qtd(x){var n=parseFloat(x); return isFinite(n)?n.toLocaleString('pt-BR'):'—';}
  function abrirModal(){ ultimoFoco=document.activeElement; modal.hidden=false;
    document.body.style.overflow='hidden'; document.getElementById('modalX').focus(); }
  function fecharModal(){ modal.hidden=true; document.body.style.overflow='';
    if(ultimoFoco&&ultimoFoco.focus)ultimoFoco.focus(); }
  document.getElementById('modalX').onclick=fecharModal;
  modal.addEventListener('click',function(e){ if(e.target===modal) fecharModal(); });
  document.addEventListener('keydown',function(e){
    if(e.key==='Escape'&&!modal.hidden) fecharModal(); });

  // TODO valor dinâmico vem de textContent (nunca innerHTML) p/ evitar XSS.
  function detalheItem(tr){
    var d=tr.dataset;
    modalBody.textContent='';
    var h=mk('h3',null,'Item '+(tr.cells[2].textContent.trim()||'—')); h.id='modalTit';
    modalBody.appendChild(h);
    modalBody.appendChild(mk('p','m-sub',
      (d.catmat||'sem CATMAT')+(d.cod?(' · código '+d.cod):'')));
    var om=OMNOMES[tr.cells[1].textContent.trim()];
    var ger=tr.cells[1].textContent.trim()+(om?(' · '+om.n):'');
    var g=mk('div','m-grid');
    g.appendChild(campo('Descrição', d.desc, true));
    g.appendChild(campo('Pregão', tr.cells[0].textContent.trim()));
    g.appendChild(campo('UASG gerenciadora', ger));
    g.appendChild(campo('Categoria', d.cat));
    g.appendChild(campo('Material ou serviço', d.tipo));
    g.appendChild(campo('ND sugerida', d.nd));
    g.appendChild(campo('Subitem sugerido', d.sub));
    g.appendChild(campo('Confiança da ND', d.conf));
    g.appendChild(campo('Fornecedor', d.forn, true));
    modalBody.appendChild(g);
    modalBody.appendChild(mk('div','m-sec','Ata e vigência'));
    var g2=mk('div','m-grid');
    g2.appendChild(campo('Nº da ata', d.nrata));
    g2.appendChild(campo('Situação', d.stlabel));
    g2.appendChild(campo('Início da vigência', d.ini));
    g2.appendChild(campo('Fim da vigência', d.fim));
    modalBody.appendChild(g2);
    modalBody.appendChild(mk('div','m-sec','Quantidades e valores'));
    var g3=mk('div','m-grid');
    g3.appendChild(campo('Qtd. homologada', qtd(d.qhom)));
    g3.appendChild(campo('Qtd. autorizada', qtd(d.qaut)));
    g3.appendChild(campo('Saldo', qtd(d.saldo)));
    g3.appendChild(campo('Valor unitário', money(d.vu)));
    g3.appendChild(campo('Valor total homologado', money(d.vtot)));
    g3.appendChild(campo('Capacidade (saldo × unit.)', money(d.cap)));
    modalBody.appendChild(g3);
    if((d.link||'').indexOf('http')===0){
      var a=mk('a','m-link','Abrir item no Compras.gov.br ↗');
      a.href=d.link; a.target='_blank'; a.rel='noopener';
      modalBody.appendChild(a);
    }
  }

  function detalhePregao(tr){
    var d=tr.dataset;
    modalBody.textContent='';
    var pg=tr.cells[1].textContent.trim();
    var h=mk('h3',null,'Pregão '+pg); h.id='modalTit'; modalBody.appendChild(h);
    var om=OMNOMES[d.ger], nomeOm=om?om.n:(d.omnome||'');
    var sub=mk('p','m-sub'); sub.appendChild(document.createTextNode(
      'Gerenciadora: UASG '+d.ger+(nomeOm?(' · '+nomeOm):'')+' · '));
    sub.appendChild(mk('span','m-badge '+(d.papel==='ger'?'p-ger':'p-part'),
      d.papel==='ger'?'BCMS é o gerenciador':'BCMS é participante'));
    modalBody.appendChild(sub);
    var venc=(tr.cells[0].querySelector('b')||{}).textContent||'';
    var dias=(tr.querySelector('.pl-dias')||{}).textContent||'';
    var g=mk('div','m-grid');
    g.appendChild(campo('Vencimento mais próximo', venc+(dias?(' — '+dias):'')));
    g.appendChild(campo('Papel do BCMS', d.papel==='ger'
      ?'Gerenciador (conduz o pregão)':'Participante (carona de outra OM)'));
    g.appendChild(campo('Objeto (categorias)', d.obj, true));
    g.appendChild(campo('Itens no pregão', tr.cells[4].textContent.trim()));
    g.appendChild(campo('Capacidade (saldo)',
      (tr.cells[5].querySelector('strong')||tr.cells[5]).textContent.trim()));
    g.appendChild(campo('Fornecedor(es)', d.forns, true));
    modalBody.appendChild(g);
    // mesmos itens que o planejamento conta: vig/v30 com vencimento AINDA futuro
    // (nunca vencidos — nem os que expiraram entre a coleta e hoje).
    var itensPg=[].slice.call(document.querySelectorAll('#tb tbody tr')).filter(
      function(r){var ts=parseFloat(r.dataset.fimts)||0;
        return r.dataset.key===d.key && (r.dataset.st==='vig'||r.dataset.st==='v30')
            && ts>0 && diasDe(ts)>=0;});
    itensPg.sort(function(a,b){return (parseFloat(b.dataset.cap)||0)-(parseFloat(a.dataset.cap)||0);});
    modalBody.appendChild(mk('div','m-sec','Itens deste pregão ('+itensPg.length+')'));
    if(itensPg.length){
      var wrap=mk('div','tblwrap'), t=mk('table');
      var thead=mk('thead'), htr=mk('tr');
      ['Item','Descrição','Categoria','Saldo','Vlr. unit.','Valor total','Fim vig.']
        .forEach(function(x,ix){var th=mk('th',ix>=3&&ix<=5?'num':null,x); htr.appendChild(th);});
      thead.appendChild(htr); t.appendChild(thead);
      var tb=mk('tbody');
      itensPg.forEach(function(r){
        var row=mk('tr');
        row.appendChild(mk('td','num', r.cells[2].textContent.trim()));
        row.appendChild(mk('td',null, r.dataset.desc||r.cells[3].textContent.trim()));
        row.appendChild(mk('td',null, r.dataset.cat||''));
        row.appendChild(mk('td','num', qtd(r.dataset.saldo)));
        row.appendChild(mk('td','num', money(r.dataset.vu)));
        row.appendChild(mk('td','num', money(r.dataset.cap)));
        row.appendChild(mk('td',null, r.dataset.fim||''));
        tb.appendChild(row);
      });
      t.appendChild(tb); wrap.appendChild(t); modalBody.appendChild(wrap);
    } else {
      modalBody.appendChild(mk('p','cap',
        'Os itens deste pregão não estão na lista (sem saldo ou já vencidos).'));
    }
  }

  function ligarDetalhe(sel, fn){
    var corpoT=document.querySelector(sel);
    if(!corpoT) return;
    corpoT.addEventListener('click',function(e){
      if(e.target.closest('a')) return;      // não intercepta o link ↗
      var tr=e.target.closest('tr'); if(!tr||!tr.dataset.key) return;
      fn(tr); abrirModal();
    });
    corpoT.addEventListener('keydown',function(e){
      if(e.key!=='Enter' && e.key!==' ') return;
      var tr=e.target.closest('tr'); if(!tr||!tr.dataset.key) return;
      e.preventDefault(); fn(tr); abrirModal();
    });
  }
  ligarDetalhe('#tb tbody', detalheItem);
  ligarDetalhe('#tbPlan tbody', detalhePregao);
})();
</script>
</body>
</html>
"""


# ---------------------------------------------------------------- main
def main() -> int:
    global UASG_ALVO, NOME_UNIDADE, NOME_CURTO, UASG_PRIMARIO
    local = os.environ.get("SOURCE_XLSX", "").strip()
    if local:
        print(f"[INFO] Usando arquivo local: {local}")
        with open(local, "rb") as f:
            fobj = io.BytesIO(f.read())
    else:
        fid = os.environ.get("DRIVE_FILE_ID", "").strip() or DRIVE_FILE_ID_PADRAO
        if "COLOQUE_AQUI" in fid:
            print("[ERRO] Defina DRIVE_FILE_ID (secret/variável) ou edite "
                  "DRIVE_FILE_ID_PADRAO no script.")
            return 2
        fobj = baixar_do_drive(fid)

    linhas = carregar_linhas(fobj)        # lê a planilha UMA vez

    # 1) ETL de cada OMDS (mesmas linhas, filtro por UASG). Guarda as que têm dados.
    resultados: dict = {}
    for u in UNIDADES:
        UASG_ALVO, NOME_UNIDADE, NOME_CURTO = u["uasg"], u["nome"], u["sigla"]
        mu = etl(linhas)
        if mu["itens"]:
            resultados[u["uasg"]] = mu
    if not resultados:
        print("[ERRO] Nenhuma OMDS com dados no consolidado — abortando.")
        return 3
    com_dados = set(resultados)
    # index.html tem de existir: se a unidade primária não tem dados, promove a
    # primeira (na ordem do manifesto) que tenha.
    if UASG_PRIMARIO not in com_dados:
        UASG_PRIMARIO = next(u["uasg"] for u in UNIDADES if u["uasg"] in com_dados)

    # 2) Uma página por OMDS com dados (primária = index.html; demais = <slug>.html).
    os.makedirs(SITE, exist_ok=True)
    for u in UNIDADES:
        if u["uasg"] not in com_dados:
            print(f"[AVISO] {u['sigla']} (UASG {u['uasg']}) sem dados — página não gerada.")
            continue
        UASG_ALVO, NOME_UNIDADE, NOME_CURTO = u["uasg"], u["nome"], u["sigla"]
        mu = resultados[u["uasg"]]
        # o histórico (history.json é único) acompanha só a unidade primária
        hist = atualizar_historico(mu) if u["uasg"] == UASG_PRIMARIO else []
        arq = pagina_da_unidade(u)
        with open(os.path.join(SITE, arq), "w", encoding="utf-8") as f:
            f.write(render(mu, hist, com_dados))
        print(f"[OK] {arq}: {fmt_brl(mu['cap_total'])} "
              f"({mu['n_itens']} itens, {mu['n_pregoes']} pregões)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
